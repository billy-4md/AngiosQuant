import os 
import json 
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from math import sqrt
import imageio

from scipy.ndimage import median_filter, zoom, rotate, gaussian_filter
from skimage import measure, morphology
from scipy.ndimage import label, find_objects
from skimage.measure import regionprops, label
from skimage.draw import disk
from skimage import io, morphology, color, measure
from skimage.transform import resize
from sklearn.decomposition import PCA
from csbdeep.utils import normalize
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

import utils
import segmentation_tool


"""-------------EXCELS FUNCTIONS----------"""
def project_and_analyze_labels(image_3d):
    labels_3d = label(image_3d)
    label_properties = {}
    center_img = (image_3d.shape[1]//2, image_3d.shape[2]//2)

    for region in regionprops(labels_3d):
        if region.area < 5:
            continue

        label_mask = labels_3d == region.label
        projection_2d = np.max(image_3d * label_mask, axis=0)

        circularity, eccentricity, angle = calculate_properties(projection_2d, center_img)
        label_properties[region.label] = {
            'Circularity': circularity,
            'Eccentricity': eccentricity,
            'Orientation': angle 
        }

    return label_properties

def calculate_properties(projection_2d, image_center):
    labeled_image = label(projection_2d)
    region = regionprops(labeled_image)[0]

    circularity = (4 * np.pi * region.area) / (region.perimeter ** 2) if region.perimeter else 0

    y, x = np.transpose(region.coords)
    points = np.vstack((x, y)).T - region.centroid
    pca = PCA(n_components=2)
    pca.fit(points)

    # First principal component
    principal_vector = pca.components_[0]

    # Vector from nucleus center to image center
    center_vector = np.array(image_center) - region.centroid

    # Normalize vectors
    principal_vector_norm = principal_vector / np.linalg.norm(principal_vector)
    center_vector_norm = center_vector / np.linalg.norm(center_vector)

    # Calculate angle in degrees between the two vectors
    angle_rad = np.arccos(np.clip(np.dot(principal_vector_norm, center_vector_norm), -1.0, 1.0))
    angle_deg = np.degrees(angle_rad)

    if angle_deg > 90:
        angle_deg -= 180
    elif angle_deg < -90:
        angle_deg += 180

    a = np.sqrt(pca.explained_variance_[0])
    b = np.sqrt(pca.explained_variance_[1])
    eccentricity = np.sqrt(1 - (b**2 / a**2)) if a > b else 0

    return circularity, eccentricity, angle_deg

def is_close(center1, center2, percent, z_weight):
    threshold = percent * shape_img[1] 
    distance = np.sqrt(((center1[0]*z_weight - center2[0]*z_weight)**2) + (center1[1] - center2[1])**2 + (center1[2] - center2[2])**2)
    return distance <= threshold

def generate_csv(threshold, z_weight, tag_center):
    with open(os.path.join(json_path, "proteins_name.json"), "r") as f:
        protein_names = json.load(f)

    tags = [tag for tag in tag_center.keys() if tag != "merged_img" and tag != "image1" and tag != "image2" and tag != "phalo"]
    columns = ["CellID"] + tags + ["Population", "Protein Name", "Circularity", "Eccentricity", "Orientation", "Center X", "Center Y"]
    data = []
    false_positive_cells = []

    label_img = imageio.volread(os.path.join(project_path, 'label_merged_img.tif'))
    label_properties = project_and_analyze_labels(label_img)

    for cell_id, center in enumerate(tag_center["merged_img"].values(), start=1):
        row = [cell_id]
        presence_counter = 0

        for tag in tags:
            if tag in tag_center:
                match = any(is_close(center, tag_center[tag][cell], threshold, z_weight) for cell in tag_center[tag])
                presence = 1 if match else 0
                row.append(presence)
                presence_counter += presence
            else:
                row.append(0)

        population = "Manual Check" 
        protein_name = ""
        if presence_counter <= 1 or presence_counter >= 6:
            population = "False positive"
            false_positive_cells.append(cell_id)
        elif presence_counter == 3:
            population = ''.join([tag[-1] for tag in tags if row[columns.index(tag)] == 1])
            protein_name = protein_names.get(population, "")

        # Ajouter la circularité et l'excentricité pour le label courant
        properties = label_properties.get(cell_id, {'Circularity': "None", 'Eccentricity':  "None", 'Orientation': "None"})
        row.extend([population, protein_name, properties['Circularity'], properties['Eccentricity'], properties['Orientation'], center[2], center[1]])
        data.append(row)

    df = pd.DataFrame(data, columns=columns)
    df.to_csv(os.path.join(project_path, f"cell_population_{threshold}_z_{z_weight}.csv"), index=False)
    generate_excel(df, columns, f"cell_population_{threshold}_z_{z_weight}.xlsx")

def update_label_image(false_positive_cells):
    full_img_labels = imageio.volread(os.path.join(project_path, 'label_full_img.tif')) #Check ici le nomn, pas sur !!!
    for cell_id in false_positive_cells:
        full_img_labels[full_img_labels == cell_id] = 0
    imageio.volwrite(os.path.join(project_path, 'label_full_img_updt.tif'), full_img_labels)

def generate_excel(df, columns, filename):
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    font = Font(bold=True)

    for col, value in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.fill = header_fill
        cell.font = font

    column_widths = {
        "Population": 15,
        "Protein Name": 15,
        "Circularity": 15,
        "Eccentricity": 15,
        "Orientation": 15,
        "Center X": 13,
        "Center Y": 13
    }

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == (columns.index("Population")+1):
                cell.fill = get_population_color(value)

    for column_name, width in column_widths.items():
        if column_name in df.columns:
            col_idx = df.columns.get_loc(column_name) + 1  # get_loc est zéro-indice, Excel est un indice basé sur 1
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Enregistrer le fichier
    wb.save(os.path.join(project_path, filename))

def get_population_color(value):
    color_false_positive = "FFC7CE"
    color_manual_check = "FFEB9C" 
    color_population = "C6E0B4"  
    if value == "False positive":
        return PatternFill(start_color=color_false_positive, end_color=color_false_positive, fill_type="solid")
    elif value == "Manual Check":
        return PatternFill(start_color=color_manual_check, end_color=color_manual_check, fill_type="solid")
    else: 
        return PatternFill(start_color=color_population, end_color=color_population, fill_type="solid")


"""-------------------------PROCESSING FUNCTIONS--------------------------"""
def translate_image(image, y_offset, x_offset, scaling_factor, scaling = True):
    if scaling:
        scaled_image = zoom(image, (scaling_factor, scaling_factor), order=0)
    else:
        scaled_image = image

    translated_image = np.zeros_like(scaled_image)
    
    for y in range(scaled_image.shape[0]):
        for x in range(scaled_image.shape[1]):
            new_y = y + y_offset
            new_x = x + x_offset
            
            if 0 <= new_y < scaled_image.shape[0] and 0 <= new_x < scaled_image.shape[1]:
                translated_image[new_y, new_x] = scaled_image[y, x]
                    
    return translated_image

def apply_translation_img(img, center_img, scaling_factor, rotation_angle, image2=False, axes=(1, 0)):
    y1, x1 = center_img

    x_middle, y_middle = img.shape[2]//2, img.shape[1]//2 
    x_offset_img1 = int(x_middle - x1)
    y_offset_img1 = int(y_middle - y1)

    img_mod = np.zeros((img.shape[0], img.shape[1], img.shape[2]), dtype=img.dtype)
    if image2:
        for z in range(img.shape[0]):
            img_mod[z, :, :] =  translate_image(img[z, :, :], y_offset_img1, x_offset_img1, scaling_factor, scaling=False)

    else:
        croppage_value_x = int(((img.shape[2] * scaling_factor) - img.shape[2]) / 2)
        croppage_value_y = int(((img.shape[1] * scaling_factor) - img.shape[1]) / 2)

        for z in range(img.shape[0]):
            slice_img1 = img[z, :, :]
            slice_img1 = translate_image(slice_img1, y_offset_img1, x_offset_img1, scaling_factor)
            slice_img1 = resize(slice_img1[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x], (img.shape[1], img.shape[2]), order=0, mode='edge', anti_aliasing=False)
            rotated_img = rotate(slice_img1, rotation_angle, reshape=False, axes=axes)
            slice_img1 = resize(rotated_img, (img.shape[1], img.shape[2]), order=0, mode='edge', anti_aliasing=False)
            img_mod[z, :, :] = slice_img1

    img_mod = normalize(img_mod, 1, 99.8, axis=(0,1,2))
    img_mod = median_filter(img_mod, size=3)
    img_mod = gaussian_filter(img_mod, sigma=1)

    return img_mod

def apply_translation_img_phalo(img, center_img, scaling_factor, rotation_angle, cropped_value, axes=(1, 0)):
    y1, x1 = center_img
    # processed_size = (img.shape[1]-(2*int(cropped_value)), img.shape[2]-(2*int(cropped_value)))
    # y1, x1 = int((y1/processed_size[0])*img.shape[1]), int((x1/processed_size[1])*img.shape[2])
    x_middle, y_middle = img.shape[2]//2, img.shape[1]//2
    x_offset_img1 = int(x_middle - y1)
    y_offset_img1 = int(y_middle - x1)
    # scaling_factor=(scaling_factor/processed_size[0])*img.shape[1]
    print(scaling_factor)

    img_mod = np.zeros_like(img)  # Utiliser zeros_like pour conserver le dtype
    croppage_value_x = int(((img.shape[2] * scaling_factor) - img.shape[2]) / 2)
    croppage_value_y = int(((img.shape[1] * scaling_factor) - img.shape[1]) / 2)

    for z in range(img.shape[0]):
        slice_img = img[z, :, :]
        slice_img = translate_image(slice_img, y_offset_img1, x_offset_img1, scaling_factor)
        slice_img = slice_img[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x]
        #slice_img = resize(slice_img, (img.shape[1], img.shape[2]), order=0, mode='edge', anti_aliasing=False)
        rotated_img = rotate(slice_img, rotation_angle, reshape=False, axes=axes)
        img_mod[z, :, :] = rotated_img 

    # Normaliser et appliquer des filtres si nécessaire
    img_mod = normalize(img_mod, 1, 99.8, axis=(0,1,2))
    img_mod = median_filter(img_mod, size=3)
    img_mod = gaussian_filter(img_mod, sigma=1)

    return img_mod

def process_image(image, project_info, tags_info, tag_center):
    global shape_img

    cropped_value = tags_info["croppingValue"]
    up_slice_to_remove1 = tags_info["image1"]["upperLayersToRemove"]
    down_slice_to_remove1 = tags_info["image1"]["lowerLayersToRemove"]
    up_slice_to_remove2 = tags_info["image2"]["upperLayersToRemove"]
    down_slice_to_remove2 = tags_info["image2"]["lowerLayersToRemove"]
    phaloidin_tag = tags_info["phaloTag"]

    scaling_factor = float(project_info["scale_factor"])
    rotation_angle = float(project_info["rotation_factor"])
    center_bead1_str1 = project_info["center_bead1"]
    center_img1 = tuple(int(float(coord)) for coord in center_bead1_str1.split(','))
    center_bead1_str2 = project_info["center_bead2"]
    center_img2 = tuple(int(float(coord)) for coord in center_bead1_str2.split(','))

    for tag_label, wavelength in tags_info[image]['tags'].items():
        if tag_label != f"TAG{phaloidin_tag}":
            tag_img = utils.isolate_and_normalize_channel(wavelength, image, tags_info)
            if image == "image1":
                tag_img_cropped = tag_img[up_slice_to_remove1:-down_slice_to_remove1, cropped_value:-cropped_value, cropped_value:-cropped_value]
                shape_img = tag_img_cropped.shape
                tag_img_cropped = apply_translation_img(tag_img_cropped, center_img1, scaling_factor, rotation_angle)

            elif image == "image2":
                tag_img_cropped = tag_img[up_slice_to_remove2:-down_slice_to_remove2, cropped_value:-cropped_value, cropped_value:-cropped_value]
                tag_img_cropped = apply_translation_img(tag_img_cropped, center_img2, scaling_factor, rotation_angle, image2=True)

            labels = segmentation_tool.do_segmentation_CellPose(tag_img_cropped, cellpose_max_th, cellpose_min_th, cellpose_diam_value)
            tag_center[tag_label] = segmentation_tool.get_center_of_labels(labels)


"""-------------------------PHALLOIDIN FUNCTIONS--------------------------"""
def isolate_and_segment_phalo(image, project_info, tags_info, tag_center):
    global shape_img

    cropped_value = tags_info["croppingValue"]
    up_slice_to_remove1 = tags_info["image1"]["upperLayersToRemove"]
    down_slice_to_remove1 = tags_info["image1"]["lowerLayersToRemove"]
    phaloidin_tag = tags_info["phaloTag"]

    scaling_factor = float(project_info["scale_factor"])
    rotation_angle = float(project_info["rotation_factor"])
    center_bead1_str = project_info["center_bead1"]
    center_img1 = tuple(int(float(coord)) for coord in center_bead1_str.split(','))

    for tag_label, wavelength in tags_info[image]['tags'].items():
        if tag_label == f"TAG{phaloidin_tag}":
            tag_img_full = utils.isolate_and_normalize_channel(wavelength, image, tags_info)
            tag_img_cropped = tag_img_full[up_slice_to_remove1:-down_slice_to_remove1, cropped_value:-cropped_value, cropped_value:-cropped_value]
            tag_img_cropped = apply_translation_img_phalo(tag_img_cropped, center_img1, scaling_factor, rotation_angle, cropped_value)
            utils.save_image_function(tag_img_cropped, "phalloidin_cropped.tif", project_path)
            utils.save_image_function(tag_img_full, "phalloidin_original.tif", project_path)
            labels = segmentation_tool.do_segmentation_CellPose_phalo(tag_img_full, cellpose_max_th_phalo, cellpose_min_th_phalo, cellpose_diam_value_phalo)
            #tag_center["phalo"] = segmentation_tool.get_center_of_labels(labels) #Uncomment if finding the center of each segmentation in the phalloidin is necessary
            label_cropped = labels[up_slice_to_remove1:-down_slice_to_remove1, cropped_value:-cropped_value, cropped_value:-cropped_value]
            label_cropped = apply_translation_img_phalo(label_cropped, center_img1, scaling_factor, rotation_angle, cropped_value)
            utils.save_image_function(label_cropped, "label_phalloidin_cropped.tif", project_path)
            utils.save_image_function(labels, "label_phalloidin_original.tif", project_path)

def check_center_distance(center, project_info):
    radius = float(project_info["radius_bead2"][0])
    center_img = (shape_img[1]//2, shape_img[1]//2)  
    distance = sqrt((center[1] - center_img[0])**2 + (center[2] - center_img[1])**2)
    threshold = 0

    if distance > (radius - radius * threshold):
        print(distance)
        return True
    else:
        return False

def match_phalo_nuclei(tag_center, project_info, z_weight):
    merged_center = tag_center["merged_img"]
    phalo_center = tag_center["phalo"]
    match_dico = {}

    for label_nuc, center_nuc in merged_center.items():
        if check_center_distance(center_nuc, project_info):
            min_distance = float('inf') 
            label_value = None  
            for label_phalo, center_phalo in phalo_center.items():
                distance = sqrt(((center_nuc[0] - center_phalo[0]) * z_weight)**2 + (center_nuc[1] - center_phalo[1])**2 + (center_nuc[2] - center_phalo[2])**2)
                if distance < min_distance:
                    min_distance = distance
                    label_value = label_phalo

            if label_value is not None:  
                match_dico[label_nuc] = label_value
            
    return match_dico

def filter_labels(label_img, labels_to_keep):
    mask = np.isin(label_img, labels_to_keep)
    filtered_img = np.where(mask, label_img, 0)
    
    return filtered_img

def keep_dist_nuclei(label_img, tag_center, project_info):
    merged_center = tag_center["merged_img"]
    label_to_keep= []

    for label_nuc, center_nuc in merged_center.items():
        if check_center_distance(center_nuc, project_info):
            label_to_keep.append(int(label_nuc))

    new_img = filter_labels(label_img, label_to_keep)
    utils.save_image_function(new_img, "label_dist_nuclei.tif", project_path)
    return label_to_keep
    

"""-------------------------PCA--------------------------"""
def calculate_circularity(area, perimeter):
    """Calcule la circularité d'une région."""
    if perimeter == 0:
        return 0
    return (4 * np.pi * area) / (perimeter ** 2)

def project_label(image_3d, label_value):
    """Projette un label spécifique en 2D en utilisant la projection maximale."""
    masked = np.where(image_3d == label_value, image_3d, 0)
    projection_2d = np.max(masked, axis=0)
    return projection_2d

def get_principal_components_2D(labels_2d_image, center_bead):
    labels = np.unique(labels_2d_image)[1:]  # ignorer le fond (label 0)
    principal_components = {}
    bead_to_label_vectors = {}
    eccentricities = {}

    for label in labels:
        y, x = np.where(labels_2d_image == label)
        points = np.vstack((x, y)).T

        if points.shape[0] < 2:
            continue

        center_label = np.mean(points, axis=0)
        points_centered = points - center_label

        pca = PCA(n_components=2)
        pca.fit(points_centered)

        vector_bead_to_label = center_label - center_bead
        vector_principal = pca.components_[0]
        norm_principal = vector_principal / np.linalg.norm(vector_principal)
        norm_bead_to_label = vector_bead_to_label / np.linalg.norm(vector_bead_to_label)
        alignment = np.dot(norm_principal, norm_bead_to_label)

        a = np.sqrt(pca.explained_variance_[0])
        b = np.sqrt(pca.explained_variance_[1])
        eccentricity = np.sqrt(1 - (b**2 / a**2)) if a > b else 0

        principal_components[label] = (pca.components_[0], pca.components_[1])
        bead_to_label_vectors[label] = {'vector': vector_bead_to_label, 'alignment': alignment}
        eccentricities[label] = eccentricity

    return principal_components, bead_to_label_vectors, eccentricities

def visualize_all(image_3d, center_bead):
    fig, ax = plt.subplots(figsize=(10, 10))  # Ajuster la taille selon vos besoins

    # Obtenir et visualiser la projection totale pour la visualisation finale
    total_projection_2d = np.max(image_3d, axis=0)
    ax.imshow(total_projection_2d, cmap='gray', alpha=1)  # Utiliser une transparence pour mieux voir les superpositions

    # Traiter chaque label individuellement et superposer les résultats
    unique_labels = np.unique(image_3d)[1:]  # Exclude background
    for label in unique_labels:
        projection_2d = project_label(image_3d, label)
        principal_components, bead_to_label_vectors, excen = get_principal_components_2D(projection_2d, center_bead)
        visualize_principal_components(projection_2d, principal_components, bead_to_label_vectors, excen, center_bead, ax)

    plt.show()

def visualize_principal_components(image_2d, principal_components, bead_to_label_vectors, eccentricities, center_bead, ax):
    for lbl, vectors in principal_components.items():
        y, x = np.where(image_2d == lbl)
        if len(x) == 0 or len(y) == 0:
            continue
        center_x = np.mean(x)
        center_y = np.mean(y)

        # Draw principal components
        vector_scaled_1 = vectors[0] * 20
        vector_scaled_2 = vectors[1] * 20
        ax.arrow(center_x, center_y, vector_scaled_1[0], vector_scaled_1[1], head_width=5, head_length=6, fc='red', ec='red')
        ax.arrow(center_x, center_y, vector_scaled_2[0], vector_scaled_2[1], head_width=5, head_length=6, fc='blue', ec='blue')

        # Calculate and display circularity
        region = regionprops(label(image_2d == lbl))[0]
        circularity = calculate_circularity(region.area, region.perimeter)
        eccentricity = eccentricities[lbl]
        ax.text(center_x, center_y + 10, f'Circ: {circularity:.2f}, Ecc: {eccentricity:.2f}', color='yellow', fontsize=12)

    ax.set_title('Individual Labels Analysis with PCA, Circularity, and Eccentricity')
    ax.axis('off')


"""-------------------------FULL IMG--------------------------"""
def create_full_img_with_phalo(project_info):
    cropped_value = int(project_info["cropping_value"])

    merged_img_path = os.path.join(project_path, "merged_img.tif")
    phallo_img_path = os.path.join(project_path, "phalloidin.tif")

    merged_img = imageio.volread(merged_img_path)
    phallo_img = imageio.volread(phallo_img_path)

    z_dim, y_dim, x_dim = min(merged_img.shape[0], phallo_img.shape[0]), merged_img.shape[1], merged_img.shape[2]
    new_img = np.zeros((z_dim, y_dim, x_dim), dtype=merged_img.dtype)

    if merged_img.shape[0] < phallo_img.shape[0]:
        diff_slice = phallo_img.shape[0] - merged_img.shape[0]
        phallo_img = phallo_img[diff_slice//2:-diff_slice//2, :, :]
    elif merged_img.shape[0] > phallo_img.shape[0]:
        diff_slice = merged_img.shape[0] - phallo_img.shape[0]
        merged_img = merged_img[diff_slice//2:-diff_slice//2, :, :]

    for z in range(z_dim):
        slice_img1 = merged_img[z, :, :]
        slice_img2 = phallo_img[z, :, :]

        new_img[z, :, :] = np.maximum(slice_img1, slice_img2)

    combined_img_path = os.path.join(project_path, "complete_img.tif")
    imageio.volwrite(combined_img_path, new_img)
    print("Image combined and saved to", combined_img_path)
    return


    # Determine the new target dimensions based on merged_img
    target_height = merged_img.shape[1]
    target_width = merged_img.shape[2]

    # Resize phallo_img to match the dimensions of merged_img
    #phallo_img_resized = resize(phallo_img, merged_img.shape, order=0, mode='edge', anti_aliasing=True)
    #padded_merged_img = np.pad(merged_img, ((0, 0), (cropped_value, cropped_value), (cropped_value, cropped_value)), mode='constant', constant_values=0)
    #merged_img_resized = resize(merged_img, phallo_img.shape, order=0, mode='edge', anti_aliasing=True)

    # Combine images
    full_img_list = [merged_img, phallo_img]
    complete_img = np.maximum.reduce(full_img_list)

    # Save the combined image
    combined_img_path = os.path.join(project_path, "complete_img.tif")
    imageio.volwrite(combined_img_path, complete_img)
    print("Image combined and saved to", combined_img_path)

    


"""-------------------------MAIN--------------------------"""
def main_find_pop(current_project):
    global json_path, project_path
    global cellpose_max_th, cellpose_min_th, cellpose_diam_value, cellpose_max_th_phalo, cellpose_min_th_phalo, cellpose_diam_value_phalo

    global shape_img
    #Global variable
    resources_path = os.environ.get('FLASK_RESOURCES_PATH', os.getcwd())
    json_path = os.path.join(resources_path, 'python', 'server', 'json') 
    project_path = os.path.join(resources_path, 'python', 'server', 'project', current_project)
    tag_center = {}
    images_list = ["image1", "image2"]
    cellpose_max_th = 10000
    cellpose_min_th = 1000
    cellpose_diam_value = 60
    cellpose_max_th_phalo = 100000
    cellpose_min_th_phalo = 100
    cellpose_diam_value_phalo = 100

    #Load json data
    json_file_path = os.path.join(json_path, 'settings_data.json')
    with open(json_file_path, 'r') as f:
        tags_info = json.load(f) 
    utils.modify_running_process("Find population process running", json_path)

    json_file_path = os.path.join(project_path, 'project_info.json')
    with open(json_file_path, 'r') as f:
        project_info = json.load(f) 

    if not tags_info["configuration"]["useTagCenter"]:
        #Each TAG will be segmented individually and finding centers
        for image in images_list:
            process_image(image, project_info, tags_info, tag_center)

        #Merged image finding centers
        merged_img = imageio.volread(os.path.join(project_path, "merged_img.tif"))
        label = segmentation_tool.do_segmentation_CellPose(merged_img, cellpose_max_th, cellpose_min_th, cellpose_diam_value)
        utils.save_image_function(label, "label_merged_img.tif", project_path)
        tag_center["merged_img"] = segmentation_tool.get_center_of_labels(label)

        #Phalloidin segmentation
        for image in images_list:
            isolate_and_segment_phalo(image, project_info, tags_info, tag_center, merged_img.shape)

    else:
        json_file_path = os.path.join(project_path, 'tag_centers.json')
        with open(json_file_path, 'r') as f:
            tag_center = json.load(f) 

    for image in images_list:
        isolate_and_segment_phalo(image, project_info, tags_info, tag_center)
    create_full_img_with_phalo(project_info)
    print("done here")
    return

    #Merged image finding centers
    # merged_img = imageio.volread(os.path.join(project_path, "merged_img.tif"))
    # label = imageio.volread(os.path.join(project_path, "label_merged_img.tif"))
    # center_bead = (624, 624)
    # visualize_all(label, center_bead)
    # plt.show()
    #max_label_projection(label)
    #imageio.imwrite(os.path.join(project_path, "proj2D.tif"), max_label_projection(label))


    
    # #tag_center["merged_img"] = segmentation_tool.get_center_of_labels(label_img)
    # label_img = imageio.volread(os.path.join(project_path, "label_merged_img.tif"))
    # shape_img = label_img.shape
    # keep_dist_nuclei(label_img, tag_center, project_info)
    # label = imageio.volread(os.path.join(project_path, "updt_merge_label.tif"))
    # center_bead = (624, 624)
    # visualize_all(label, center_bead)
    # plt.show()
    # # utils.save_tag_centers(tag_center, project_path)
    # #match_phalo_dico = match_phalo_nuclei(tag_center, project_info, 0.75)
    # #print(match_phalo_dico)
            

    #return #TO remove, the code below is the one to generatee the excel file.
    #Generate csv
    label_img = imageio.volread(os.path.join(project_path, "label_merged_img.tif"))
    shape_img = label_img.shape
    threshold_distance = project_info.get("threshold_distance_pop", 0.05) #Set default value
    z_weight = project_info.get("z_weight_pop", 0.75)  #Set default value
    threshold_distance, z_weight = float(threshold_distance), float(z_weight) 
    generate_csv(threshold_distance, z_weight, tag_center)

    #Complete program
    project_info["threshold_distance_pop"] = f"{threshold_distance}"
    project_info["z_weight_pop"] = f"{z_weight}"
    project_info["last_merging_run"] = "Matching population"
    file_path = os.path.join(project_path, "project_info.json")
    data_project_prepared = utils.prepare_data_for_json(project_info)
    with open(file_path, 'w') as file:
        json.dump(data_project_prepared, file, indent=4)

    utils.save_tag_centers(tag_center, project_path)
    utils.modify_running_process("Done population process", json_path)

    # except Exception as e:
    #     utils.modify_running_process("Error while running the matching population process", json_path)
    #     print(e)

