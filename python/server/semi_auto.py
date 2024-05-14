import os
import sys
import json
import imageio
import czifile
import re
import cv2
import numpy as np
import pandas as pd


from xml.etree import ElementTree as ET
from csbdeep.utils import normalize
from scipy.ndimage import median_filter, center_of_mass, zoom, shift, rotate, affine_transform, gaussian_filter
from scipy.ndimage import label, find_objects
from skimage.morphology import binary_closing, binary_opening, ball
from skimage.transform import resize
from scipy.spatial.distance import cdist
from sklearn.decomposition import PCA
from stardist.models import StarDist2D, StarDist3D
from skimage.morphology import disk
from skimage.filters import threshold_otsu
from mpl_toolkits.mplot3d import Axes3D

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

import open3d as o3d
from scipy.spatial.transform import Rotation as R
import matplotlib.pyplot as plt


"""-------------CZI IMAGES FUNCTIONS----------"""
def open_image(image_path):
    with czifile.CziFile(image_path) as czi:
        image_czi = czi.asarray()
        dic_dim = dict_shape(czi)  
        axes = czi.axes  
        metadata_xml = czi.metadata()
        metadata = ET.fromstring(metadata_xml)
        channels_dict = channels_dict_def(metadata)

    return image_czi, dic_dim, channels_dict, axes

#return a dictionary of the czi shape dimensions:
def dict_shape(czi):
    return dict(zip([*czi.axes],czi.shape))

def channels_dict_def(metadata):
    channels = metadata.findall('.//Channel')
    channels_dict = {}
    for chan in channels:
        name = chan.attrib['Name']
        dye_element = chan.find('DyeName')
        if dye_element is not None:
            dye_numbers = re.findall(r'\d+', dye_element.text)
            dye_number = dye_numbers[-1] if dye_numbers else 'Unknown'
            channels_dict[dye_number] = name
    return channels_dict

def get_channel_index(channel_name, channel_dict):
    count = 0
    for keys, values in channel_dict.items():
        if values == channel_name:
            return count
        else:
            count += 1
  
def czi_slicer(arr,axes,indexes={"S":0, "C":0}):
    ret = arr
    axes = [*axes]
    for k,v in indexes.items():
        index = axes.index(k)
        axes.remove(k)
        ret = ret.take(axis=index, indices=v)
 
    ret_axes = ""
    for i,v in enumerate(ret.shape):
        if v > 1:
            ret_axes+=axes[i]
    return ret.squeeze(),ret_axes
 
def isolate_and_normalize_channel(TAG, image_name):
    image, dic_dim, channel_dict, axes = open_image(tags_info[image_name]['path'])
    channel_name = channel_dict[TAG]
    channel_index = get_channel_index(channel_name, channel_dict)
    if channel_index < 0 or channel_index >= dic_dim['C']:
        raise ValueError("Channel index out of range.")

    image_czi_reduite, axes_restants = czi_slicer(image, axes, indexes={'C': channel_index})

    return image_czi_reduite.astype(np.uint16)


"""-------------HOUGH FUNCTIONS----------"""
def find_circles(composite_image, name):  
    layer_to_keep = 4    
    segmentation_thresholds = (1000000, 10)

    mid_layer = composite_image.shape[0] // 2
    focused_image = composite_image[mid_layer-layer_to_keep:mid_layer+layer_to_keep, :, :]

    labels_all_slices, data = do_segmentation(focused_image, *segmentation_thresholds)
    labels_all_slices = np.stack(labels_all_slices, axis=0)
    df = pd.DataFrame(data)
    labels = reassign_labels(labels_all_slices, df)
    #save_image_function(labels, f'label_full_{name}.tif')

    img_2D_full = project_3d_to_2d_min_layers(labels)
    circles = hough_transform(img_2D_full)
    print(circles) #To remove - ONLY for debug
    return circles

def project_3d_to_2d_min_layers(image_3d, min_layers=2, median_size=5, closing_size=3):
    """
    Projette une image 3D en 2D en ne conservant que les pixels présents dans un minimum de couches,
    avec traitement par filtre médian et fermeture morphologique pour améliorer la qualité.

    Args:
    - image_3d (numpy.ndarray): Image 3D d'entrée.
    - min_layers (int): Le nombre minimum de couches où un pixel doit être présent pour être conservé.
    - median_size (int): Taille du voisinage pour le filtre médian.
    - closing_size (int): Taille de l'élément structurant pour la fermeture morphologique.

    Returns:
    - numpy.ndarray: Image 2D projetée après traitement.
    """
    # Count how many layers each pixel is present in
    label_count = np.count_nonzero(image_3d > 0, axis=0)
    mask = label_count >= min_layers
    image_2d = np.zeros((image_3d.shape[1], image_3d.shape[2]), dtype=image_3d.dtype)

    # Apply projection logic
    for z in range(image_3d.shape[0]):
        update_mask = (image_2d == 0) & mask & (image_3d[z] > 0)
        image_2d[update_mask] = image_3d[z][update_mask]

    # Apply a median filter to reduce noise
    image_2d = median_filter(image_2d, size=median_size)

    # Apply morphological closing to fill gaps and connect nearby objects
    #selem = disk(closing_size)  # Create a disk-shaped structural element
    #image_2d = binary_closing(image_2d, selem)

    # Save the processed image
    #imageio.imwrite(os.path.join(project_path, "img_2d.tif"), image_2d.astype(np.uint8))

    return image_2d
    
def hough_transform(image):
    image = np.uint8(image)
    circles = cv2.HoughCircles(image, cv2.HOUGH_GRADIENT, dp=2, minDist=200, param1=10, param2=50, minRadius=250, maxRadius=450)
    return circles



"""-------------MERGING FUNCTIONS----------"""
def translate_image(image, y_offset, x_offset, scaling_factor, scaling = True):
    if scaling:
        scaled_image = zoom(image, (scaling_factor, scaling_factor), order=0)
    else:
        scaled_image = image

    translated_image = np.zeros_like(scaled_image)
    
    for y in range(scaled_image.shape[0]):
        for x in range(scaled_image.shape[1]):
            new_y = y + y_offset
            new_x = x + x_offset
            
            if 0 <= new_y < scaled_image.shape[0] and 0 <= new_x < scaled_image.shape[1]:
                translated_image[new_y, new_x] = scaled_image[y, x]
                    
    return translated_image

def merge_images(img1, img2, angle, axes=(1, 0)):

    z_dim, y_dim, x_dim = min(img1.shape[0], img2.shape[0]), img1.shape[1], img1.shape[2]
    merged_img = np.zeros((z_dim, y_dim, x_dim), dtype=img1.dtype)

    if img1.shape[0] < img2.shape[0]:
        diff_slice = img2.shape[0] - img1.shape[0]
        img2 = img2[diff_slice//2:-diff_slice//2, :, :]
    elif img1.shape[0] > img2.shape[0]:
        diff_slice = img1.shape[0] - img2.shape[0]
        img1 = img1[diff_slice//2:-diff_slice//2, :, :]


    for z in range(z_dim):
        slice_img1 = img1[z, :, :]
        slice_img2 = img2[z, :, :]

        rotated_img = rotate(slice_img1, angle, reshape=False, axes=axes)
        slice_img1 = resize(rotated_img, (y_dim, x_dim), order=0, mode='edge', anti_aliasing=False)

        merged_img[z, :, :] = np.maximum(slice_img1, slice_img2)

    save_image_function(merged_img, 'img_merged.tif')

    return merged_img

def apply_transfo_img(img1, img2, center_img1, center_img2, scaling_factor, image_to_save=True):
    y1, x1 = center_img1
    y2, x2 = center_img2

    x_middle, y_middle = img1.shape[2]//2, img1.shape[1]//2 #Center of the image to translate
    x_offset_img1, x_offset_img2  = int(x_middle - x1), int(x_middle - x2)
    y_offset_img1, y_offset_img2 = int(y_middle - y1), int(y_middle - y2)

    img1_mod = np.zeros((img1.shape[0], img1.shape[1], img1.shape[2]), dtype=img1.dtype)
    img2_mod = np.zeros((img2.shape[0], img2.shape[1], img2.shape[2]), dtype=img2.dtype)

    croppage_value_x = int(((img1.shape[2] * scaling_factor) - img1.shape[2]) / 2)
    croppage_value_y = int(((img1.shape[1] * scaling_factor) - img1.shape[1]) / 2)

    for z in range(img1.shape[0]):
        slice_img1 = img1[z, :, :]
        slice_img1 = translate_image(slice_img1, y_offset_img1, x_offset_img1, scaling_factor)
        slice_img1 = resize(slice_img1[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x], (img1.shape[1], img1.shape[2]), order=0, mode='edge', anti_aliasing=False)
        img1_mod[z, :, :] = slice_img1

    for z in range(img2.shape[0]):
        slice_img2 = img2[z, :, :]
        slice_img2 = translate_image(slice_img2, y_offset_img2, x_offset_img2, scaling_factor, scaling=False) #Scaling = False allow to only make the translation and not the scaling
        img2_mod[z, :, :] = slice_img2

    if image_to_save:
        save_image_function(img1_mod, f'image1_processed.tif')
        save_image_function(img2_mod, f'image2_processed.tif')

    return img1_mod, img2_mod

def apply_transfo_on_label(label_img1, label_img2, center_img1, center_img2, scaling_factor, angle):
    """
    Applique une mise à l'échelle et une rotation aux images étiquetées.
    Effectue une rotation autour de l'axe Z en faisant pivoter chaque tranche.
    """
    label_img1_mod, label_img2_mod = apply_transfo_img(label_img1, label_img2, center_img1, center_img2, scaling_factor, image_to_save=False)

    # Faire pivoter chaque tranche 2D dans le plan XY
    for z in range(label_img1_mod.shape[0]):
        label_img1_mod[z, :, :] = rotate(label_img1_mod[z, :, :], angle, reshape=False, mode='constant', cval=0)

    # Sauvegarder les images transformées
    save_image_function(label_img1_mod, "label_image1_processed.tif")
    save_image_function(label_img2_mod, "label_image2_processed.tif")

    return label_img1_mod, label_img2_mod

def apply_translation_img(img, center_img, scaling_factor, rotation_angle, image2=False, axes=(1, 0)):
    y1, x1 = center_img

    x_middle, y_middle = img.shape[2]//2, img.shape[1]//2 #Center of the image to translate
    x_offset_img1 = int(x_middle - x1)
    y_offset_img1 = int(y_middle - y1)

    img_mod = np.zeros((img.shape[0], img.shape[1], img.shape[2]), dtype=img.dtype)
    if image2:
        for z in range(img.shape[0]):
            img_mod[z, :, :] =  translate_image(img[z, :, :], y_offset_img1, x_offset_img1, scaling_factor, scaling=False)

    else:
        croppage_value_x = int(((img.shape[2] * scaling_factor) - img.shape[2]) / 2)
        croppage_value_y = int(((img.shape[1] * scaling_factor) - img.shape[1]) / 2)

        for z in range(img.shape[0]):
            slice_img1 = img[z, :, :]
            slice_img1 = translate_image(slice_img1, y_offset_img1, x_offset_img1, scaling_factor)
            slice_img1 = resize(slice_img1[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x], (img.shape[1], img.shape[2]), order=0, mode='edge', anti_aliasing=False)
            rotated_img = rotate(slice_img1, rotation_angle, reshape=False, axes=axes)
            slice_img1 = resize(rotated_img, (img.shape[1], img.shape[2]), order=0, mode='edge', anti_aliasing=False)
            img_mod[z, :, :] = slice_img1

    img_mod = normalize(img_mod, 1, 99.8, axis=(0,1,2))
    img_mod = median_filter(img_mod, size=3)
    img_mod = gaussian_filter(img_mod, sigma=1)

    return img_mod


"""-------------VECTORS FUNCTIONS----------"""
def compare_vectors(vectors1, vectors2, match):
    """
    Compare les vecteurs 2 à 2 entre les images brutes et transformées.
    """
    differences = {}
    count = 1

    for label1, label2 in match.items():
        vec1 = vectors1[label1]
        vec2 = vectors2[label2]
        # Comparer les vecteurs par la distance euclidienne ou l'angle
        difference = np.linalg.norm(vec1 - vec2)

        # Normaliser le produit scalaire pour qu'il soit compris entre -1 et 1
        norm_vec1 = vec1 / np.linalg.norm(vec1) if np.linalg.norm(vec1) != 0 else vec1
        norm_vec2 = vec2 / np.linalg.norm(vec2) if np.linalg.norm(vec2) != 0 else vec2

        # Calculer le produit scalaire entre les vecteurs normalisés
        dot_product = np.dot(norm_vec1, norm_vec2)
        
        # Limiter le produit scalaire entre -1 et 1 pour éviter les erreurs arccos
        dot_product = max(min(dot_product, 1.0), -1.0)
        
        # Trouver l'angle en radians, puis le convertir en degrés
        angle_radians = np.arccos(dot_product)
        angle_degrees = np.degrees(angle_radians)

        differences[count] = {
            "label1": label1,
            "label2": label2,
            "distance_difference": difference,
            "angle_difference": angle_degrees
        }

        count += 1

    return differences

def load_control_points(file_path):
    """
    Charge les points de contrôle à partir d'un fichier .txt.
    """
    control_points = []
    with open(file_path, 'r') as file:
        for line in file:
            point = tuple(map(float, line.strip().split(',')))
            control_points.append(point)
    return np.array(control_points)

def find_closest_labels(centers, control_points, z_weight=0.5):
    """
    Trouve le numéro de label du centre de nucléus le plus proche de chaque point de contrôle.
    """
    matched_labels = []

    for point in control_points:
        z, y, x = point
        distances = []
        
        for label, (cz, cy, cx) in centers.items():
            distance = np.sqrt((z_weight * (z - cz))**2 + (y - cy)**2 + (x - cx)**2)
            distances.append((distance, label))
        
        closest_label = min(distances, key=lambda t: t[0])[1]
        matched_labels.append(closest_label)
    
    return matched_labels

def calculate_vectors(label_list, center_bille, image_name, z_dim):
    """
    Calcule les vecteurs de chaque centre par rapport au centre de la bille.
    """
    vectors = {}
    for label in label_list:
        z, y, x = tag_center[image_name][label]
        vector = np.array([z - z_dim, y - center_bille[0], x - center_bille[1]])
        #normalized_vector = vector / np.linalg.norm(vector)
        vectors[label] = vector
    return vectors

def find_vector_matches(vectors1, vectors2, threshold=0.95):
    """
    Trouve les correspondances entre les vecteurs similaires dans les deux ensembles.
    La correspondance est basée sur le produit scalaire entre les vecteurs.
    """
    matches = {}
    for label1, vec1 in vectors1.items():
        best_match = None
        best_similarity = -1
        
        for label2, vec2 in vectors2.items():
            similarity = np.dot(vec1, vec2)
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = label2
        
        if best_similarity >= threshold:
            matches[label1] = best_match
    
    return matches

def calculate_scaling_factor(matches, center_bille1, center_bille2, z_dim1, z_dim2):
    """
    Calcule le facteur d'échelle entre deux images en utilisant les distances par rapport aux centres de la bille.
    """
    scaling_factors = []
    
    for label1, label2 in matches.items():
        z1, y1, x1 = tag_center["image1"][label1]
        z2, y2, x2 = tag_center["image2"][label2]
        
        vec1 = np.array([z1 - z_dim1, y1 - center_bille1[0], x1 - center_bille1[1]])
        vec2 = np.array([z2 - z_dim2, y2 - center_bille2[0], x2 - center_bille2[1]])
        
        distance1 = np.linalg.norm(vec1)
        distance2 = np.linalg.norm(vec2)
        
        if distance1 != 0:
            scaling_factor = distance2 / distance1
            scaling_factors.append(scaling_factor)
    
    scaling_factor = np.mean(scaling_factors)
    print(f"Scaling factor: {scaling_factor}")
    return scaling_factor

def calculate_angle(v):
    """
    Calcule l'angle en radians du vecteur donné par rapport à l'axe des x.
    """
    return np.arctan2(v[1], v[0])

def calculate_rotation_angle(matches, center_bille1, center_bille2):
    """
    Calcule l'angle de rotation moyen entre deux ensembles de points en utilisant leurs correspondances
    et leurs positions relatives au centre de la bille.
    """
    rotation_angles = []
    
    for label1, label2 in matches.items():
        z1, y1, x1 = tag_center["image1"][label1]
        z2, y2, x2 = tag_center["image2"][label2]
        
        vec1 = np.array([y1 - center_bille1[0], x1 - center_bille1[1]])
        vec2 = np.array([y2 - center_bille2[0], x2 - center_bille2[1]])
        
        norm_vec1 = vec1 / np.linalg.norm(vec1) if np.linalg.norm(vec1) != 0 else vec1
        norm_vec2 = vec2 / np.linalg.norm(vec2) if np.linalg.norm(vec2) != 0 else vec2
        
        angle1 = calculate_angle(norm_vec1)
        angle2 = calculate_angle(norm_vec2)
        
        angle_difference = angle2 - angle1
        rotation_angles.append(angle_difference)
    
    average_angle = np.degrees(np.mean(rotation_angles))
    print(f"Rotation angle: {average_angle}")

    return average_angle
    


"""-------------PLOT FUNCTIONS----------"""
def plot_vectors(label_list, center_bille, vectors, image_name):
    """
    Affiche les vecteurs relatifs au centre de la bille sur un graphique.
    """
    plt.figure(figsize=(8, 8))
    plt.scatter(center_bille[1], center_bille[0], color='red', label='Bead center')
    
    for label in label_list:
        z, y, x = tag_center[image_name][label]
        vec = vectors[label]
        plt.arrow(x, y, vec[1] * 50, vec[0] * 50, head_width=10, head_length=10, fc='blue', ec='blue')
        plt.text(x, y, label, fontsize=12)
    
    plt.gca().invert_yaxis() 
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.title(f"Vectors {image_name}")
    plt.legend()
    plt.grid(True)
    plt.show()

def plot_comparison_vectors(center_bille1, vectors1, label_list1, center_bille2, vectors2, label_list2):
    """
    Affiche les vecteurs relatifs au centre de la bille pour deux ensembles sur un seul graphique.
    """
    plt.figure(figsize=(10, 10))
    print(label_list1)
    print(label_list2)
    
    # Afficher les vecteurs pour la première image
    plt.scatter(center_bille1[1], center_bille1[0], color='red', label='Bead center (Image 1)')
    for label in label_list1:
        z, y, x = tag_center["image1"][label]
        vec = vectors1[label]
        plt.arrow(x, y, vec[1] * 50, vec[0] * 50, head_width=10, head_length=10, fc='blue', ec='blue')
        #plt.text(x, y, label, fontsize=12)
    
    # Afficher les vecteurs pour la deuxième image
    plt.scatter(center_bille2[1], center_bille2[0], color='green', label='Bead center (Image 2)')
    for label in label_list2:
        z, y, x = tag_center["image2"][label]
        vec = vectors2[label]
        plt.arrow(x, y, vec[1] * 50, vec[0] * 50, head_width=10, head_length=10, fc='orange', ec='orange')
        #plt.text(x, y, label, fontsize=12)

    plt.plot([], [], color='blue', label='Vectors (Image 1)')
    plt.plot([], [], color='orange', label='Vectors (Image 2)')

    plt.gca().invert_yaxis() 
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.legend()
    plt.grid(True)
    plt.title('Vectors between 2 images')
    plt.show()

def plot_comparison_vectors_3d(center_bille1, vectors1, label_list1, center_bille2, vectors2, label_list2, z_dim1, z_dim2):
    """
    Affiche les vecteurs relatifs au centre de la bille pour deux ensembles dans un graphique 3D.
    """
    fig = plt.figure(figsize=(12, 8))
    ax = fig.add_subplot(111, projection='3d')

    # Afficher le centre de la bille pour la première image
    ax.scatter(z_dim1, center_bille1[1], center_bille1[0], color='red', label='Bead Center (Image 1)')

    for label in label_list1:
        z, y, x = tag_center["image1"][label]
        vec = vectors1[label]
        ax.quiver(x, y, z, vec[2], vec[1], vec[0], length=50, color='blue', label='Vectors (Image 1)' if label == label_list1[0] else '')

    # Afficher le centre de la bille pour la deuxième image
    ax.scatter(z_dim2, center_bille2[1], center_bille2[0], color='green', label='Bead Center (Image 2)')

    for label in label_list2:
        z, y, x = tag_center["image2"][label]
        vec = vectors2[label]
        ax.quiver(x, y, z, vec[2], vec[1], vec[0], length=50, color='orange', label='Vectors (Image 2)' if label == label_list2[0] else '')

    ax.set_xlabel('X')
    ax.set_ylabel('Y')
    ax.set_zlabel('Z')
    ax.legend()
    ax.grid(True)
    ax.set_title('Vectors between 2 Images (3D)')
    plt.show()



"""-------------GENERAL FUNCTIONS----------"""
def get_normalized_center_of_labels(label_image):
    unique_labels = np.unique(label_image)[1:]
    dims = np.array(label_image.shape)

    normalized_centers = {}

    for label in unique_labels:
        center = np.array(center_of_mass(label_image == label))
        normalized_center = center / dims
        normalized_centers[label] = normalized_center

    return normalized_centers

def get_center_of_labels(label_image):
    unique_labels = np.unique(label_image)[1:]

    centers = {}

    for label in unique_labels:
        center = np.array(center_of_mass(label_image == label))
        centers[label] = center

    return centers

def create_full_image(image_name):
    full_img = []

    for tag_label, wavelength in tags_info[image_name]['tags'].items():
        tag_img = isolate_and_normalize_channel(wavelength, image_name)
        if image_name == "image1":
            tag_img_cropped = tag_img[up_slice_to_remove1:-down_slice_to_remove1, cropped_value:-cropped_value, cropped_value:-cropped_value]
        elif image_name == "image2":
            tag_img_cropped = tag_img[up_slice_to_remove2:-down_slice_to_remove2, cropped_value:-cropped_value, cropped_value:-cropped_value]
        
        tag_img_cropped = normalize(tag_img_cropped, 1, 99.8, axis=(0,1,2))
        tag_img_cropped = median_filter(tag_img_cropped, size=3)
        gaussian_slice = gaussian_filter(tag_img_cropped, sigma=1)

        full_img.append(gaussian_slice)

    complete_img = np.maximum.reduce(full_img)
    save_image_function(complete_img, f"full_{image_name}.tif")
    return complete_img

def modify_running_process(value):
    if value == "ON":
        tags_info["process"] = "Yes"
    else:
        tags_info["process"] = "None"
    json_file_path = os.path.join(json_path, 'settings_data.json')
    with open(json_file_path, 'w') as file:
        json.dump(tags_info, file, indent=4)  

def save_image_function(image, image_name):
    if save_images:
        imageio.volwrite(os.path.join(project_path, image_name), image)
        #imageio.volwrite(os.path.join("D:\\Users\\MFE\\Xavier\\Result_Rapport2_new", image_name), image)


"""-------------EXCELS FUNCTIONS----------"""
def is_close(center1, center2, percent, z_weight):
    threshold = percent * shape_img 
    distance = np.sqrt(((center1[0]*z_weight - center2[0]*z_weight)**2) + (center1[1] - center2[1])**2 + (center1[2] - center2[2])**2)
    return distance <= threshold

def generate_csv(th, z_w):
    with open(os.path.join(json_path, "proteins_name.json"), "r") as f:
        protein_names = json.load(f)

    tags = [tag for tag in tag_center.keys() if tag != "full_img" and tag != "image1" and tag != "image2"]
    columns = ["CellID"] + tags + ["Population", "Protein Name", "Center X", "Center Y"]
    data = []
    false_positive_cells = []

    for cell_id, center in enumerate(tag_center["full_img"].values(), start=1):
        row = [cell_id]
        presence_counter = 0

        for tag in tags:
            if tag in tag_center:
                match = any(is_close(center, tag_center[tag][cell], th, z_w) for cell in tag_center[tag])
                presence = 1 if match else 0
                row.append(presence)
                presence_counter += presence
            else:
                row.append(0)

        if presence_counter <= 1 or presence_counter >= 6:
            population = "False positive"
            protein_name = ""
            false_positive_cells.append(cell_id)
        elif presence_counter == 3:
            population = ''.join([tag[-1] for tag in tags if row[columns.index(tag)] == 1])
            protein_name = protein_names.get(population, "")
        else:
            population = "Manual Check"
            protein_name = ""

        row.extend([population, protein_name, center[2], center[1]]) #DON'T FARGET TO MULTIPLY BY THE DIM OF THE IMAGE TO NOT HAVE THE NORMALIZED CENTER!!!!!!!!
        data.append(row)

    df = pd.DataFrame(data, columns=columns)
    df.to_csv(os.path.join(project_path, f"cell_population_{th}_z_{z_w}.csv"), index=False)

    update_label_image(false_positive_cells)

    generate_excel(df, columns, th, z_w)

def update_label_image(false_positive_cells):
    full_img_labels = imageio.volread(os.path.join(project_path, 'label_full_img.tif'))
    for cell_id in false_positive_cells:
        full_img_labels[full_img_labels == cell_id] = 0
    imageio.volwrite(os.path.join(project_path, 'label_full_img_updt.tif'), full_img_labels)

def generate_excel(df, columns, th, z_w):
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    font = Font(bold=True)

    for col, value in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.fill = header_fill
        cell.font = font

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Configuration des couleurs de fond pour les cellules de la colonne "Population"
            if c_idx == (columns.index("Population")+1):
                cell.fill = get_population_color(value)

    wb.save(os.path.join(project_path, f"cell_population_{th}_z_{z_w}.xlsx"))

def get_population_color(value):
    color_false_positive = "FFC7CE"
    color_manual_check = "FFEB9C" 
    color_population = "C6E0B4"  
    if value == "False positive":
        return PatternFill(start_color=color_false_positive, end_color=color_false_positive, fill_type="solid")
    elif value == "Manual Check":
        return PatternFill(start_color=color_manual_check, end_color=color_manual_check, fill_type="solid")
    else: 
        return PatternFill(start_color=color_population, end_color=color_population, fill_type="solid")



"""-------------SEGMENTATION PROCESS FUNCTIONS----------"""
def do_segmentation(isolate_image, SURFACE_THRESHOLD_SUP, SURFACE_THRESHOLD_INF):
    labels_all_slices = []
    data = []

    for z in range(isolate_image.shape[0]):
        img_slice = isolate_image[z, :, :]
        img_slice = normalize(img_slice, 1, 99.8)
        #filtered_img_slice = median_filter(img_slice, size=9)
        filtered_img_slice = img_slice 
        labels, details = model.predict_instances(filtered_img_slice)
        unique_labels = np.unique(labels)

        labels_to_remove = []
        for label_num in unique_labels:
            if label_num == 0: 
                continue

            instance_mask = labels == label_num
            label_surface = np.sum(instance_mask)

            if label_surface > SURFACE_THRESHOLD_SUP or label_surface < SURFACE_THRESHOLD_INF:
                labels[instance_mask] = 0
                labels_to_remove.append(label_num)

        unique_labels = np.array([label for label in unique_labels if label not in labels_to_remove])
        for label in unique_labels:
            if label == 0:  
                continue
            center = center_of_mass(labels == label)
            data.append({'Layer': z, 'Label': label, 'Center X': center[0], 'Center Y': center[1]})

        labels_all_slices.append(labels)

    return labels_all_slices, data

def reassign_labels(labels_all_slices, df):
    new_labels_all_slices = []
    max_label = 0  # Pour garder une trace du dernier label utilisé
    seuil = np.mean(labels_all_slices[0].shape) * 0.02

    for z in range(len(labels_all_slices)):
        labels = labels_all_slices[z]
        new_labels = np.zeros_like(labels)
        layer_df = df[df['Layer'] == z]

        # Obtenir les dataframes pour les couches précédente et suivante, si elles existent
        prev_layer_df = df[df['Layer'] == z - 1] if z > 0 else None
        next_layer_df = df[df['Layer'] == z + 1] if z < len(labels_all_slices) - 1 else None

        for idx, row in layer_df.iterrows():
            label = row['Label']
            if label == 0:
                continue
            
            # Trouver les labels similaires dans les couches précédente et suivante
            current_center = np.array([row['Center X'], row['Center Y']])
            similar_label_found = False

            for adjacent_layer_df in [prev_layer_df, next_layer_df]:
                if adjacent_layer_df is not None:
                    adj_centers = adjacent_layer_df[['Center X', 'Center Y']].values
                    if len(adj_centers) > 0:
                        distances = cdist([current_center], adj_centers)
                        min_dist_idx = np.argmin(distances)
                        min_dist = distances[0, min_dist_idx]

                        if min_dist < seuil:
                            similar_label_found = True
                            adj_label = adjacent_layer_df.iloc[min_dist_idx]['Label']
                            new_labels[labels == label] = new_labels_all_slices[z - 1][labels_all_slices[z - 1] == adj_label][0] if adjacent_layer_df is prev_layer_df else max_label + 1
                            break
            if similar_label_found:
                max_label += 1

        new_labels_all_slices.append(new_labels)

    return reassign_labels_sequentially(np.stack(new_labels_all_slices, axis=0))

def reassign_labels_sequentially(labels):
    unique_labels = np.unique(labels)
    unique_labels = unique_labels[unique_labels != 0]  # Exclure le label 0 s'il représente le fond
    
    # Créer un dictionnaire pour mapper les anciens labels vers les nouveaux
    label_mapping = {old_label: new_label for new_label, old_label in enumerate(unique_labels, start=1)}
    
    # Appliquer le mappage pour réassigner les labels
    new_labels = np.copy(labels)
    for old_label, new_label in label_mapping.items():
        new_labels[labels == old_label] = new_label
    
    return new_labels

def unify_labels(image1, image2, match_dico):
    """
    Applique les mêmes labels entre deux images segmentées en fonction du dictionnaire `match_dico`.

    Args:
    - image1: Première image segmentée avec des labels numériques.
    - image2: Deuxième image segmentée avec des labels numériques.
    - match_dico: Dictionnaire de correspondance entre les labels de `image1` et `image2`.

    Returns:
    - image1_sync: Première image synchronisée avec les nouveaux labels.
    - image2_sync: Deuxième image synchronisée avec les nouveaux labels.
    """
    # Trouver tous les labels uniques dans les deux images
    image1_sync = np.zeros_like(image1)
    image2_sync = np.zeros_like(image2)

    for label1, label2 in match_dico.items():
        new_label = label1

        image1_sync[image1 == label1] = new_label
        image2_sync[image2 == label2] = new_label

    return image1_sync, image2_sync

def create_2d_projection(label_volume):
    """
    Creates a 2D projection from a 3D volume by accumulating the presence of each label across the Z-axis.
    
    Args:
    - label_volume (numpy.ndarray): The 3D numpy array of labeled segments.
    
    Returns:
    - numpy.ndarray: A 2D array where each entry is a set of labels that appear in that position across all Z layers.
    """
    z_dim, y_dim, x_dim = label_volume.shape
    projection = np.zeros((y_dim, x_dim), dtype=object)

    for y in range(y_dim):
        for x in range(x_dim):
            labels = set(label_volume[:, y, x])
            if 0 in labels:
                labels.remove(0)  # Remove background label if present
            projection[y, x] = labels

    return projection

def unify_labels_overlap(label_volume):
    """
    Unifies labels in the 3D volume based on their overlap detected in the 2D projection.
    
    Args:
    - label_volume (numpy.ndarray): The original 3D label volume.
    - projection (numpy.ndarray): The 2D projection with sets of labels at each position.
    
    Returns:
    - numpy.ndarray: The updated 3D label volume with unified labels.
    """
    label_map = {}
    projection = create_2d_projection(label_volume)
    y_dim, x_dim = projection.shape

    # Create a mapping of labels to unify them
    for y in range(y_dim):
        for x in range(x_dim):
            if len(projection[y, x]) > 1:
                sorted_labels = sorted(projection[y, x])
                base_label = sorted_labels[0]
                for label in sorted_labels[1:]:
                    label_map[label] = base_label

    # Update the 3D volume according to the label map
    unified_label_volume = np.copy(label_volume)
    for label in label_map:
        unified_label_volume[label_volume == label] = label_map[label]

    return unified_label_volume


"""-------------GENERAL PROCESS FUNCTIONS----------"""
def process_image(image, center_img1, center_img2, scaling_factor, rotation_angle):
    for tag_label, wavelength in tags_info[image]['tags'].items():
        tag_img = isolate_and_normalize_channel(wavelength, image)
        if image == "image1":
            tag_img_cropped = tag_img[up_slice_to_remove1:-down_slice_to_remove1, cropped_value:-cropped_value, cropped_value:-cropped_value]
            tag_img_cropped = apply_translation_img(tag_img_cropped, center_img1, scaling_factor, rotation_angle)

        elif image == "image2":
            tag_img_cropped = tag_img[up_slice_to_remove2:-down_slice_to_remove2, cropped_value:-cropped_value, cropped_value:-cropped_value]
            tag_img_cropped = apply_translation_img(tag_img_cropped, center_img2, scaling_factor, rotation_angle, image2=True)

        save_image_function(tag_img_cropped, f'{tag_label}_processed.tif')
        start_segmentation(tag_img_cropped, tag_label, 5000, 1500)
        
def start_segmentation(image, image_name, area_sup, area_inf):
    labels_all_slices, data = do_segmentation(image, SURFACE_THRESHOLD_SUP = area_sup, SURFACE_THRESHOLD_INF = area_inf)

    labels_all_slices = np.stack(labels_all_slices, axis=0)
    df = pd.DataFrame(data)

    labels = reassign_labels(labels_all_slices, df)
    labels = unify_labels_overlap(labels)
    save_image_function(labels, f'label_{image_name}.tif')

    tag_center[image_name] = get_center_of_labels(labels)

    return labels

def stringify_keys(data):
    """
    Convertit récursivement toutes les clés de dictionnaire en chaînes de caractères.
    Gère également les listes contenant des dictionnaires.
    """
    if isinstance(data, dict):
        return {str(key): stringify_keys(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [stringify_keys(item) for item in data]
    else:
        return data

def convert_numpy(obj):
    if isinstance(obj, np.ndarray):
        return obj.tolist()  # Convertit les arrays numpy en listes
    elif isinstance(obj, dict):
        return {key: convert_numpy(value) for key, value in obj.items()}  # Appel récursif pour les dictionnaires
    elif isinstance(obj, list):
        return [convert_numpy(item) for item in obj]  # Appel récursif pour les listes
    else:
        return obj
    


"""-------------MAIN----------"""
def launch_segmentation_process():
    global json_path, project_path
    global tag_center, tags_info, model
    global count_pop, cropped_value, up_slice_to_remove1, up_slice_to_remove2, down_slice_to_remove1, down_slice_to_remove2
    global save_images
    global image1_composite, image2_composite
    global shape_img

    #Global variable
    resources_path = os.environ.get('FLASK_RESOURCES_PATH', os.getcwd())
    json_path = os.path.join(resources_path, 'python', 'server', 'json') 

    #Load json data
    json_file_path = os.path.join(json_path, 'settings_data.json')
    with open(json_file_path, 'r') as f:
        tags_info = json.load(f)
    #Modify the json file (To put the process running inside the application)   
    modify_running_process("ON")
    

    project_name = tags_info["project_name"]
    project_path = os.path.join(resources_path, 'python', 'server', 'project', project_name) 
    project_path = "D:\\Users\\MFE\\Xavier\\Semi_auto" #----------------------------TO CHANGE!!!!!!!!!!!!!!!!!!!!! (REMOVE)--------------------------
    model = StarDist2D.from_pretrained('2D_versatile_fluo') #Load StarDist model 
    tag_center = {}
    tag_PCA = {}
    image1 = []
    image2 = []
    count_pop = 1 
    cropped_value = tags_info["croppingValue"]
    up_slice_to_remove1 = tags_info["image1"]["upperLayersToRemove"]
    down_slice_to_remove1 = tags_info["image1"]["lowerLayersToRemove"]
    up_slice_to_remove2 = tags_info["image2"]["upperLayersToRemove"]
    down_slice_to_remove2 = tags_info["image2"]["lowerLayersToRemove"]
    save_images = tags_info["configuration"]["saveAllImages"]
    images = ["image1", "image2"]
    
    #Configuration first
    if not os.path.exists(project_path):
        os.mkdir(project_path)

    if True:
        file_path = os.path.join(project_path, "tag_centers.json")
        with open(file_path, 'r') as file:
            tag_center = json.load(file)

        file_path = os.path.join(project_path, "project_info.json")
        with open(file_path, 'r') as file:
            data_project = json.load(file)

        shape_img =(int(data_project['center_img'][0]))*2        
            
        value_threshold = [0.01, 0.02, 0.03, 0.04, 0.05]
        z_weight = [0.1, 0.25, 0.5, 0.75, 1]
        #Generate CSV
        for z_w in z_weight:
            for th in value_threshold:
                print(f"Start generating csv {th} and z {z_w}...")
                generate_csv(th, z_w)
                print(f"Done generating csv {th} and z {z_w}")

    else:

        #Creation of needed variables/images/...
        print("HERE")
        #image1_composite = create_full_image("image1") #Create a full image by combining all the TAG in 1 image - Control points will be on those images
        #image2_composite = create_full_image("image2")
        image1_composite = imageio.volread("D:\\Users\\MFE\\Xavier\\Semi_auto\\full_image1.tif")
        image2_composite = imageio.volread("D:\\Users\\MFE\\Xavier\\Semi_auto\\full_image2.tif")

        #1st: Find center with Hough Transform
        print("HERE2")
        #center_img1 = find_circles(image1_composite, "image1")[0][0][:2] #Coordinates of the center of the bead (only y and x)
        #center_img2 = find_circles(image2_composite, "image2")[0][0][:2]
        center_img1 = (643, 587)
        center_img2 = (629, 571)

        #2nd: Apply full segmentation on each image and get normalized center
        print("HERE3")
        # full_label_img1 = start_segmentation(image1_composite, "image1", 4500, 1500)
        # full_label_img2 = start_segmentation(image2_composite, "image2", 5000, 1500)
        full_label_img1 = imageio.volread("D:\\Users\\MFE\\Xavier\\Semi_auto\\label_image1.tif")
        tag_center["image1"] = get_center_of_labels(full_label_img1)
        full_label_img2 = imageio.volread("D:\\Users\\MFE\\Xavier\\Semi_auto\\label_image2.tif")
        tag_center["image2"] = get_center_of_labels(full_label_img2)

        #3rd: Find matching point between control points and nuclei center for each image
        print("HERE4")
        control_points_path1 = os.path.join(project_path, "control_point1.txt") 
        control_points_path2 = os.path.join(project_path, "control_point2.txt")  
        control_points1 = load_control_points(control_points_path1)
        control_points2 = load_control_points(control_points_path2)
        closest_labels1 = find_closest_labels(tag_center["image1"], control_points1) #List with the label values of the matching nucleus/control point
        closest_labels2 = find_closest_labels(tag_center["image2"], control_points2)
        

        #4th: Calculate vector for each nucleus image and find corresponding vector to match the nucleus between the 2 images
        print("HERE5")
        z_dim1 = image1_composite.shape[0]
        z_dim2 = image2_composite.shape[0]
        vectors_image1 = calculate_vectors(closest_labels1, center_img1, "image1", z_dim1)
        vectors_image2 = calculate_vectors(closest_labels2, center_img2, 'image2', z_dim2)
        matches = find_vector_matches(vectors_image1, vectors_image2)
        match_label_img1, match_label_img2 = unify_labels(full_label_img1, full_label_img2, matches)
        save_image_function(match_label_img1, "match_label_img1.tif")
        save_image_function(match_label_img2, "match_label_img2.tif")

        #plot_comparison_vectors(center_img1, vectors_image1, closest_labels1, center_img2, vectors_image2, closest_labels2)
        #plot_comparison_vectors_3d(center_img1, vectors_image1, closest_labels1, center_img2, vectors_image2, closest_labels2, z_dim1, z_dim2)

        #plot_vectors(closest_labels1, center_img1, vectors_image1, 'image1')
        #plot_vectors(closest_labels2, center_img2, vectors_image2, 'image2')

        #5th: Calculate the scaling between the 2 images
        print("HERE6")
        scaling_factor = calculate_scaling_factor(matches, center_img1, center_img2, z_dim1, z_dim2)

        #6th: Apply translation + scaling factor on the original image
        print("HERE7")
        img1_mod, img2_mod = apply_transfo_img(image1_composite, image2_composite, center_img1, center_img2, scaling_factor)

        #7th: Find rotation angle and apply 
        print("HERE8")
        rotation_angle = calculate_rotation_angle(matches, center_img1, center_img2)

        #8th: Apply rotation and final merging image
        print("HERE9")
        merged_img = merge_images(img1_mod, img2_mod, rotation_angle)  

        #9th: Apply transformation on label image
        print("HERE10")
        label1_processed, label2_processed = apply_transfo_on_label(full_label_img1, full_label_img2, center_img1, center_img2, scaling_factor, rotation_angle)
        tag_center["image1"] = get_center_of_labels(label1_processed)
        tag_center["image2"] = get_center_of_labels(label2_processed)
        x_middle, y_middle = image1_composite.shape[2]//2, image1_composite.shape[1]//2
        center_img1_new = (y_middle, x_middle)
        center_img2_new = (y_middle, x_middle)
        z_dim = min(image1_composite.shape[0], image2_composite.shape[0])
        vectors_image1_new = calculate_vectors(closest_labels1, center_img1_new, "image1", z_dim)
        vectors_image2_new = calculate_vectors(closest_labels2, center_img2_new, 'image2', z_dim)
        

        #plot_comparison_vectors(center_img1, vectors_image1_new, closest_labels1, center_img2, vectors_image2_new, closest_labels2)
        #plot_comparison_vectors_3d(center_img1, vectors_image1_new, closest_labels1, center_img2, vectors_image2_new, closest_labels2, z_dim, z_dim)
        

        #10th: Analyse comparaison between vectors
        # diff = compare_vectors(vectors_image1, vectors_image2, matches)
        # for label, diff in diff.items():
        #     print(f"Label1: {diff['label1']} and Label2:  {diff['label2']}  ")
        #     print(f"  Différence de distance : {diff['distance_difference']:.4f}")
        #     print(f"  Différence d'angle (en degrés) : {diff['angle_difference']:.2f}")

        # diff_new = compare_vectors(vectors_image1_new, vectors_image2_new, matches)
        # for label, diff in diff_new.items():
        #     print(f"Label1: {diff['label1']} and Label2:  {diff['label2']}  ")
        #     print(f"  Différence de distance : {diff['distance_difference']:.4f}")
        #     print(f"  Différence d'angle (en degrés) : {diff['angle_difference']:.2f}")


        #Images Processing
        print("ICI")
        for image in images:
            process_image(image, center_img1, center_img2, scaling_factor, rotation_angle)
        print("ICI2")

        label_merged = start_segmentation(merged_img, "full_img", 5000, 1500)
        save_image_function(label_merged, "label_full_img.tif")

        #Save all centers in a json file
        file_path = os.path.join(project_path, "tag_centers.json")
        tag_center_str_keys = stringify_keys(tag_center)
        tag_center_str_keys_arr = convert_numpy(tag_center_str_keys)
        with open(file_path, 'w') as file:
            json.dump(tag_center_str_keys_arr, file, indent=4)
        print(f"Dictionary saved to {file_path}")

        

        #Generate CSV
        print("Start generating csv...")
        generate_csv(merged_img.shape)
        print("Done generating csv")


        #Save project info in a json file
        data_project = {
            "scale_factor": scaling_factor,
            "rotation_factor": rotation_angle,
            "center_bead": [center_img1, center_img2],
            "center_img": (image1_composite.shape[1]//2, image1_composite.shape[2]//2),
            "img1_processing": [up_slice_to_remove1, down_slice_to_remove1],
            "img2_processing": [up_slice_to_remove2, down_slice_to_remove2],
            "cropping_value": cropped_value       
        }

        file_path = os.path.join(project_path, "project_info.json")
        with open(file_path, 'w') as file:
            json.dump(data_project, file, indent=4)

launch_segmentation_process()



