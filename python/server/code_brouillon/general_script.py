import os
import sys
import json
import imageio
import czifile
import re
import cv2
import numpy as np
import pandas as pd


from xml.etree import ElementTree as ET
from csbdeep.utils import normalize
from scipy.ndimage import median_filter, center_of_mass, zoom, shift, rotate, affine_transform, gaussian_filter
from skimage.morphology import binary_closing, binary_opening, ball
from skimage.transform import resize
from scipy.spatial.distance import cdist
from sklearn.decomposition import PCA
from stardist.models import StarDist2D, StarDist3D
from skimage.morphology import disk
from skimage.filters import threshold_otsu

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

import open3d as o3d
from scipy.spatial.transform import Rotation as R


"""-------------CZI IMAGES FUNCTIONS----------"""
def open_image(image_path):
    with czifile.CziFile(image_path) as czi:
        image_czi = czi.asarray()
        dic_dim = dict_shape(czi)  
        axes = czi.axes  
        metadata_xml = czi.metadata()
        metadata = ET.fromstring(metadata_xml)
        channels_dict = channels_dict_def(metadata)

    return image_czi, dic_dim, channels_dict, axes

#return a dictionary of the czi shape dimensions:
def dict_shape(czi):
    return dict(zip([*czi.axes],czi.shape))

def channels_dict_def(metadata):
    channels = metadata.findall('.//Channel')
    channels_dict = {}
    for chan in channels:
        name = chan.attrib['Name']
        dye_element = chan.find('DyeName')
        if dye_element is not None:
            dye_numbers = re.findall(r'\d+', dye_element.text)
            dye_number = dye_numbers[-1] if dye_numbers else 'Unknown'
            channels_dict[dye_number] = name
    return channels_dict

def get_channel_index(channel_name, channel_dict):
    count = 0
    for keys, values in channel_dict.items():
        if values == channel_name:
            return count
        else:
            count += 1
  
def czi_slicer(arr,axes,indexes={"S":0, "C":0}):
    ret = arr
    axes = [*axes]
    for k,v in indexes.items():
        index = axes.index(k)
        axes.remove(k)
        ret = ret.take(axis=index, indices=v)
 
    ret_axes = ""
    for i,v in enumerate(ret.shape):
        if v > 1:
            ret_axes+=axes[i]
    return ret.squeeze(),ret_axes
 
def isolate_and_normalize_channel(TAG, image_name):
    image, dic_dim, channel_dict, axes = open_image(tags_info[image_name]['path'])
    channel_name = channel_dict[TAG]
    channel_index = get_channel_index(channel_name, channel_dict)
    if channel_index < 0 or channel_index >= dic_dim['C']:
        raise ValueError("Channel index out of range.")

    image_czi_reduite, axes_restants = czi_slicer(image, axes, indexes={'C': channel_index})

    return image_czi_reduite.astype(np.uint16)


"""-------------HOUGH FUNCTIONS----------"""
def find_circles(image_name):  
    layer_to_keep = 4    
    segmentation_thresholds = (1000000, 10)

    composite_image = create_full_image(image_name)
    save_image_function(composite_image, f'full_{image_name}.tif')

    mid_layer = composite_image.shape[0] // 2
    print(mid_layer)
    focused_image = composite_image[mid_layer-layer_to_keep:mid_layer+layer_to_keep, :, :]

    labels_all_slices, data = do_segmentation(focused_image, *segmentation_thresholds)
    labels_all_slices = np.stack(labels_all_slices, axis=0)
    df = pd.DataFrame(data)
    labels = reassign_labels(labels_all_slices, df)
    save_image_function(labels, f'label_full_{image_name}.tif')

    img_2D_full = project_3d_to_2d_min_layers(labels)
    circles = hough_transform(img_2D_full)
    print(circles)
    return circles

def project_3d_to_2d_min_layers(image_3d, min_layers=2, median_size=5, closing_size=3):
    """
    Projette une image 3D en 2D en ne conservant que les pixels présents dans un minimum de couches,
    avec traitement par filtre médian et fermeture morphologique pour améliorer la qualité.

    Args:
    - image_3d (numpy.ndarray): Image 3D d'entrée.
    - min_layers (int): Le nombre minimum de couches où un pixel doit être présent pour être conservé.
    - median_size (int): Taille du voisinage pour le filtre médian.
    - closing_size (int): Taille de l'élément structurant pour la fermeture morphologique.

    Returns:
    - numpy.ndarray: Image 2D projetée après traitement.
    """
    # Count how many layers each pixel is present in
    label_count = np.count_nonzero(image_3d > 0, axis=0)
    mask = label_count >= min_layers
    image_2d = np.zeros((image_3d.shape[1], image_3d.shape[2]), dtype=image_3d.dtype)

    # Apply projection logic
    for z in range(image_3d.shape[0]):
        update_mask = (image_2d == 0) & mask & (image_3d[z] > 0)
        image_2d[update_mask] = image_3d[z][update_mask]

    # Apply a median filter to reduce noise
    image_2d = median_filter(image_2d, size=median_size)

    # Apply morphological closing to fill gaps and connect nearby objects
    #selem = disk(closing_size)  # Create a disk-shaped structural element
    #image_2d = binary_closing(image_2d, selem)

    # Save the processed image
    imageio.imwrite(os.path.join(project_path, "img_2d.tif"), image_2d.astype(np.uint8))

    return image_2d
    
def hough_transform(image):
    image = np.uint8(image)
    circles = cv2.HoughCircles(image, cv2.HOUGH_GRADIENT, dp=2, minDist=200, param1=10, param2=50, minRadius=250, maxRadius=450)
    return circles

def calcul_scale_factor():
    radius1 = circles1[0][0][2]
    radius2 = circles2[0][0][2]
    scale_factor = radius2 / radius1 if radius1 > 0 else 1
    print(f"Scale factor: {scale_factor}")
    sys.stdout.flush()
    return scale_factor


"""-------------MERGING FUNCTIONS----------"""
def interpolate_layers(image_3d):
    """
    Interpole de nouvelles couches dans une image 3D pour atteindre une profondeur spécifiée.
    
    Args:
    - image_3d (numpy.ndarray): Image 3D d'entrée (profondeur, hauteur, largeur).
    
    Returns:
    - numpy.ndarray: Image 3D avec le nombre de couches ajusté par interpolation.
    """

    # Les axes y et x (hauteur et largeur) restent inchangés (zoom_factor de 1)
    interpolated_image = zoom(image_3d, (scale_factor, 1, 1), order=3)  # order=3 pour une interpolation spline cubique

    return interpolated_image

def translate_image(image, y_offset, x_offset, value = True):
    if value:
        scaled_image = zoom(image, (scale_factor, scale_factor), order=0)
    else:
        scaled_image = image

    translated_image = np.zeros_like(scaled_image)
    
    for y in range(scaled_image.shape[0]):
        for x in range(scaled_image.shape[1]):
            new_y = y + y_offset
            new_x = x + x_offset
            
            # Vérifie si les nouvelles coordonnées sont dans les limites de l'image
            if 0 <= new_y < scaled_image.shape[0] and 0 <= new_x < scaled_image.shape[1]:
                translated_image[new_y, new_x] = scaled_image[y, x]
                    
    return translated_image

def merge_images(angle, axes=(1, 0)):
    img1 = imageio.volread(os.path.join(project_path, "image1_processed.tif"))
    img2 = imageio.volread(os.path.join(project_path, "image2_processed.tif"))

    z_dim, y_dim, x_dim = min(img1.shape[0], img2.shape[0]), img1.shape[1], img1.shape[2]
    merged_img_pos = np.zeros((z_dim, y_dim, x_dim), dtype=img1.dtype)
    merged_img_neg = np.zeros((z_dim, y_dim, x_dim), dtype=img1.dtype)

    if img1.shape[0] < img2.shape[0]:
        diff_slice = img2.shape[0] - img1.shape[0]
        img2 = img2[diff_slice//2:-diff_slice//2, :, :]
    elif img1.shape[0] > img2.shape[0]:
        diff_slice = img1.shape[0] - img2.shape[0]
        img1 = img1[diff_slice//2:-diff_slice//2, :, :]


    for z in range(z_dim):
        slice_img1 = img1[z, :, :]
        slice_img2 = img2[z, :, :]

        rotated_img_pos = rotate(slice_img1, angle[0], reshape=False, axes=axes)
        rotated_img_neg = rotate(slice_img1, -angle[0], reshape=False, axes=axes)
        slice_img1_pos = resize(rotated_img_pos, (y_dim, x_dim), order=0, mode='edge', anti_aliasing=False)
        slice_img1_neg = resize(rotated_img_neg, (y_dim, x_dim), order=0, mode='edge', anti_aliasing=False)

        merged_img_pos[z, :, :] = np.maximum(slice_img1_pos, slice_img2)
        merged_img_neg[z, :, :] = np.maximum(slice_img1_neg, slice_img2)

    save_image_function(merged_img_pos, 'img_merged_pos.tif')
    save_image_function(merged_img_neg, 'img_merged_neg.tif')
    print("MERGED HERE DONE")

def apply_transfo_img(value):
    img1 = create_full_image("image1")
    img2 = create_full_image("image2")
    center1 = circles1[0][0][:2]
    center2 = circles2[0][0][:2]

    x_middle, y_middle = img1.shape[2]//2, img1.shape[1]//2 #Center of the image to translate
    x_offset_img1, x_offset_img2  = int(x_middle - center1[0]), int(x_middle - center2[0])
    y_offset_img1, y_offset_img2 = int(y_middle - center1[1]), int(y_middle - center2[1])

    img1_mod = np.zeros((img1.shape[0], img1.shape[1], img1.shape[2]), dtype=img1.dtype)
    img2_mod = np.zeros((img2.shape[0], img2.shape[1], img2.shape[2]), dtype=img2.dtype)

    if value == 1:
        croppage_value_x = int(((img1.shape[2] * scale_factor) - img1.shape[2]) / 2)
        croppage_value_y = int(((img1.shape[1] * scale_factor) - img1.shape[1]) / 2)
        for z in range(img1.shape[0]):
            slice_img1 = img1[z, :, :]
            slice_img1 = translate_image(slice_img1, y_offset_img1, x_offset_img1)
            slice_img1 = resize(slice_img1[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x], (img1.shape[1], img1.shape[2]), order=0, mode='edge', anti_aliasing=False)
            img1_mod[z, :, :] = slice_img1
        for z in range(img2.shape[0]):
            slice_img2 = img2[z, :, :]
            slice_img2 = translate_image(slice_img2, y_offset_img2, x_offset_img2, value=False) #Value = False allow to only make the translation and not the scaling
            img2_mod[z, :, :] = slice_img2
        img1_mod_v2 = interpolate_layers(img1_mod)
        save_image_function(img1_mod_v2, f'image1_processed_interpoller.tif')
        print(f"Interpoller shape: {img1_mod_v2.shape}")
        print(f"Normal shape: {img1_mod.shape}")
        print("Interpollation done")
        

    elif value == 2:
        croppage_value_x = int(((img2.shape[2] * scale_factor) - img2.shape[2]) / 2)
        croppage_value_y = int(((img2.shape[1] * scale_factor) - img2.shape[1]) / 2)
        for z in range(img1.shape[0]):
            slice_img1 = img1[z, :, :]
            slice_img1 = translate_image(slice_img1, y_offset_img1, x_offset_img1, value=False)
            img1_mod[z, :, :] = slice_img1
        for z in range(img2.shape[0]):
            slice_img2 = img2[z, :, :]
            slice_img2 = translate_image(slice_img2, y_offset_img2, x_offset_img2)
            slice_img2 = resize(slice_img2[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x], (img2.shape[1], img2.shape[2]), order=0, mode='edge', anti_aliasing=False)
            img2_mod[z, :, :] = slice_img2
        #img2_mod = interpolate_layers(img2_mod)

    save_image_function(img1_mod, f'image1_processed.tif')
    save_image_function(img2_mod, f'image2_processed.tif')
    start_segmentation(img1_mod, f'image1_processed', 4500, 1500)
    start_segmentation(img2_mod, f'image2_processed', 4500, 1500)
    keep_closest_point_between_img()

def keep_closest_point_between_img(z_weight=0.9):
    center_img1 = tag_center["image1_processed"]
    center_img2 = tag_center["image2_processed"]
    
    pt_cloud1 = []
    pt_cloud2 = []

    for (label1, center1), (label2, center2) in zip(center_img1.items(), center_img2.items()):
        weighted_center1 = np.array([center1[0], center1[1], center1[2] * z_weight])
        weighted_center2 = np.array([center2[0], center2[1], center2[2] * z_weight])
        pt_cloud1.append(weighted_center1)
        pt_cloud2.append(weighted_center2)

    print(len(pt_cloud1))
    print(len(pt_cloud2))
    #angles = start_icp(pt_cloud1, pt_cloud2, 0.05)
    angles = start_icp(pt_cloud1, pt_cloud2, 0.1)
    #angles = start_icp(pt_cloud1, pt_cloud2, 0.25)
    #angles = start_icp(pt_cloud1, pt_cloud2, 0.5)
    #angles = start_icp(pt_cloud1, pt_cloud2, 0.75)
    #angles=[-5.5, 0]
    merge_images(angles)

def prepare_point_cloud(points):
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(points)
    return pcd

def compute_fpfh_feature(pcd):
    radius_normal = 0.1
    radius_feature = 0.25
    pcd.estimate_normals(search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
    fpfh = o3d.pipelines.registration.compute_fpfh_feature(
        pcd,
        o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
    return fpfh

def perform_icp_point_to_plane(source_pcd, target_pcd, threshold):
    trans_init = np.eye(4)
    result = o3d.pipelines.registration.registration_icp(
        source_pcd, target_pcd, threshold, trans_init,
        o3d.pipelines.registration.TransformationEstimationPointToPlane())
    return result

def perform_icp_ransac(source_pcd, target_pcd, source_fpfh, target_fpfh, distance_threshold):
    result = o3d.pipelines.registration.registration_ransac_based_on_feature_matching(
        source_pcd, target_pcd, source_fpfh, target_fpfh, True,
        distance_threshold,
        o3d.pipelines.registration.TransformationEstimationPointToPoint(False),
        4,  # Four points to determine a plane
        [o3d.pipelines.registration.CorrespondenceCheckerBasedOnEdgeLength(0.9),
         o3d.pipelines.registration.CorrespondenceCheckerBasedOnDistance(distance_threshold)],
        o3d.pipelines.registration.RANSACConvergenceCriteria(1000000, 0.999))
    return result

def perform_basic_icp(source_pcd, target_pcd, threshold):
    trans_init = np.eye(4)  

    result = o3d.pipelines.registration.registration_icp(
        source_pcd, target_pcd, threshold, trans_init,
        o3d.pipelines.registration.TransformationEstimationPointToPoint())
    return result

def start_icp(points1, points2, threshold):
    pcd1 = prepare_point_cloud(points1)
    pcd2 = prepare_point_cloud(points2)
    fpfh1 = compute_fpfh_feature(pcd1)
    fpfh2 = compute_fpfh_feature(pcd2)

    print(f"--------------------Value : {threshold}---------------")

    result_basic = perform_basic_icp(pcd1, pcd2, threshold)
    print("Transformation matrix Basic:")
    print(result_basic.transformation)
    rotation_matrix = np.array(result_basic.transformation[:3, :3], dtype=np.float64, copy=True)
    rotation = R.from_matrix(rotation_matrix)
    euler_angles = rotation.as_euler('xyz', degrees=True) 
    print(euler_angles)

    # result_ransac = perform_icp_ransac(pcd1, pcd2, fpfh1, fpfh2, threshold)
    # print("Transformation matrix Ransac:")
    # print(result_ransac.transformation)
    # rotation_matrix = np.array(result_ransac.transformation[:3, :3], dtype=np.float64, copy=True)
    # rotation = R.from_matrix(rotation_matrix)
    # euler_angles = rotation.as_euler('xyz', degrees=True) 
    # print(euler_angles)

    # result_plane = perform_icp_point_to_plane(pcd1, pcd2, threshold)
    # print("Transformation matrix Point-to-Plane:")
    # print(result_plane.transformation)
    # rotation_matrix = np.array(result_plane.transformation[:3, :3], dtype=np.float64, copy=True)
    # rotation = R.from_matrix(rotation_matrix)
    # euler_angles = rotation.as_euler('xyz', degrees=True) 
    # print(euler_angles)

    return euler_angles



"""-------------GENERAL FUNCTIONS----------"""
def apply_translation_img(img, image1):
    center1 = circles1[0][0][:2]
    center2 = circles2[0][0][:2]
    img_factor = circles2[0][0][2] / circles1[0][0][2]
    croppage_value_x = int(((img.shape[2] * scale_factor) - img.shape[2]) / 2)
    croppage_value_y = int(((img.shape[1] * scale_factor) - img.shape[1]) / 2)

    z_dim, y_dim, x_dim = img.shape[0], (img.shape[1]), (img.shape[2])
    merged_img = np.zeros((z_dim, y_dim, x_dim), dtype=img.dtype)
    print(merged_img.shape)

    if img_factor > 1 and image1:               
        x_middle, y_middle = center2[0], center2[1]
        x_offset = int(x_middle - center1[0])
        y_offset = int(y_middle - center1[1])
        for z in range(z_dim):
            current_slice = translate_image(img[z, :, :], y_offset, x_offset)
            merged_img[z, :, :] = resize(current_slice[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x], (y_dim, x_dim), order=0, mode='edge', anti_aliasing=False)
        return merged_img
    
    elif img_factor < 1 and image1:
        return img

    elif img_factor > 1 and not image1:
        return img
    
    elif img_factor < 1 and not image1:
        x_middle, y_middle = center1[0], center1[1]
        x_offset = int(x_middle - center2[0])
        y_offset = int(y_middle - center2[1])
        for z in range(z_dim):
            merged_img[z, :, :] = translate_image(img[z, :, :], y_offset, x_offset)
            merged_img[z, :, :] = resize(merged_img[croppage_value_y:-croppage_value_y, croppage_value_x:-croppage_value_x], (y_dim, x_dim), order=0, mode='edge', anti_aliasing=False)
        return merged_img

def get_normalized_center_of_labels(label_image):
    unique_labels = np.unique(label_image)[1:]
    dims = np.array(label_image.shape)

    normalized_centers = {}

    for label in unique_labels:
        center = np.array(center_of_mass(label_image == label))
        normalized_center = center / dims
        normalized_centers[label] = normalized_center

    return normalized_centers

def create_full_image(image_name):
    full_img = []
    full_img_basic = []
    struct_element = ball(1)

    for tag_label, wavelength in tags_info[image_name]['tags'].items():
        tag_img = isolate_and_normalize_channel(wavelength, image_name)
        if image_name == "image1":
            tag_img_cropped = tag_img[up_slice_to_remove1:-down_slice_to_remove1, cropped_value:-cropped_value, cropped_value:-cropped_value]
        elif image_name == "image2":
            tag_img_cropped = tag_img[up_slice_to_remove2:-down_slice_to_remove2, cropped_value:-cropped_value, cropped_value:-cropped_value]
        
        full_img_basic.append(tag_img_cropped)
        tag_img_cropped = normalize(tag_img_cropped, 1, 99.8, axis=(0,1,2))
        tag_img_cropped = median_filter(tag_img_cropped, size=3)
        gaussian_slice = gaussian_filter(tag_img_cropped, sigma=1)
        full_img.append(gaussian_slice)

    save_image_function(np.maximum.reduce(full_img_basic), f"image_basic_{image_name}.tif")
    return np.maximum.reduce(full_img)

def modify_running_process(value):
    if value == "ON":
        tags_info["process"] = "Yes"
    else:
        tags_info["process"] = "None"
    json_file_path = os.path.join(json_path, 'settings_data.json')
    with open(json_file_path, 'w') as file:
        json.dump(tags_info, file, indent=4)  

def save_image_function(image, image_name):
    if save_images:
        imageio.volwrite(os.path.join(project_path, image_name), image)
        #imageio.volwrite(os.path.join("D:\\Users\\MFE\\Xavier\\Result_Rapport2_new", image_name), image)


"""-------------EXCELS FUNCTIONS----------"""
def is_close(center1, center2, threshold=0.0075, z_weight = 0.01):
    distance = np.sqrt(((center1[0]*z_weight - center2[0]*z_weight)**2) + (center1[1] - center2[1])**2 + (center1[2] - center2[2])**2)
    return distance <= threshold

def generate_csv():
    with open(os.path.join(json_path, "proteins_name.json"), "r") as f:
        protein_names = json.load(f)

    tags = [tag for tag in tag_center.keys() if tag != "full_img"]
    columns = ["CellID"] + tags + ["Population", "Protein Name", "Center X", "Center Y"]
    data = []
    false_positive_cells = []

    for cell_id, center in enumerate(tag_center["full_img"].values(), start=1):
        row = [cell_id]
        presence_counter = 0

        for tag in tags:
            if tag in tag_center:
                match = any(is_close(center, tag_center[tag][cell]) for cell in tag_center[tag])
                presence = 1 if match else 0
                row.append(presence)
                presence_counter += presence
            else:
                row.append(0)

        if presence_counter <= 1:
            population = "False positive"
            protein_name = ""
            false_positive_cells.append(cell_id)
        elif presence_counter == 3:
            population = ''.join([tag[-1] for tag in tags if row[columns.index(tag)] == 1])
            protein_name = protein_names.get(population, "")
        else:
            population = "Manual Check"
            protein_name = ""

        row.extend([population, protein_name, center[2], center[1]]) #DON'T FARGET TO MULTIPLY BY THE DIM OF THE IMAGE TO NOT HAVE THE NORMALIZED CENTER!!!!!!!!
        data.append(row)

    df = pd.DataFrame(data, columns=columns)
    df.to_csv(os.path.join(project_path, "cell_population.csv"), index=False)

    update_label_image(false_positive_cells)

    generate_excel(df, columns)

def update_label_image(false_positive_cells):
    full_img_labels = imageio.volread(os.path.join(project_path, 'label_full_img.tif'))
    for cell_id in false_positive_cells:
        full_img_labels[full_img_labels == cell_id] = 0
    imageio.volwrite(os.path.join(project_path, 'label_full_img_updt.tif'), full_img_labels)

def generate_excel(df, columns):
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    font = Font(bold=True)

    for col, value in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.fill = header_fill
        cell.font = font

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Configuration des couleurs de fond pour les cellules de la colonne "Population"
            if c_idx == (columns.index("Population")+1):
                cell.fill = get_population_color(value)

    wb.save(os.path.join(project_path, "cell_population.xlsx"))

def get_population_color(value):
    color_false_positive = "FFC7CE"
    color_manual_check = "FFEB9C" 
    color_population = "C6E0B4"  
    if value == "False positive":
        return PatternFill(start_color=color_false_positive, end_color=color_false_positive, fill_type="solid")
    elif value == "Manual Check":
        return PatternFill(start_color=color_manual_check, end_color=color_manual_check, fill_type="solid")
    else: 
        return PatternFill(start_color=color_population, end_color=color_population, fill_type="solid")


"""-------------SEGMENTATION PROCESS FUNCTIONS----------"""
def do_segmentation(isolate_image, SURFACE_THRESHOLD_SUP, SURFACE_THRESHOLD_INF):
    labels_all_slices = []
    data = []

    for z in range(isolate_image.shape[0]):
        img_slice = isolate_image[z, :, :]
        img_slice = normalize(img_slice, 1, 99.8)
        #filtered_img_slice = median_filter(img_slice, size=9)
        filtered_img_slice = img_slice 
        labels, details = model.predict_instances(filtered_img_slice)
        unique_labels = np.unique(labels)

        labels_to_remove = []
        for label_num in unique_labels:
            if label_num == 0: 
                continue

            instance_mask = labels == label_num
            label_surface = np.sum(instance_mask)

            if label_surface > SURFACE_THRESHOLD_SUP or label_surface < SURFACE_THRESHOLD_INF:
                labels[instance_mask] = 0
                labels_to_remove.append(label_num)

        unique_labels = np.array([label for label in unique_labels if label not in labels_to_remove])
        for label in unique_labels:
            if label == 0:  
                continue
            center = center_of_mass(labels == label)
            data.append({'Layer': z, 'Label': label, 'Center X': center[0], 'Center Y': center[1]})

        labels_all_slices.append(labels)

    return labels_all_slices, data

def reassign_labels(labels_all_slices, df):
    new_labels_all_slices = []
    max_label = 0  # Pour garder une trace du dernier label utilisé
    seuil = np.mean(labels_all_slices[0].shape) * 0.05

    for z in range(len(labels_all_slices)):
        labels = labels_all_slices[z]
        new_labels = np.zeros_like(labels)
        layer_df = df[df['Layer'] == z]

        # Obtenir les dataframes pour les couches précédente et suivante, si elles existent
        prev_layer_df = df[df['Layer'] == z - 1] if z > 0 else None
        next_layer_df = df[df['Layer'] == z + 1] if z < len(labels_all_slices) - 1 else None

        for idx, row in layer_df.iterrows():
            label = row['Label']
            if label == 0:
                continue
            
            # Trouver les labels similaires dans les couches précédente et suivante
            current_center = np.array([row['Center X'], row['Center Y']])
            similar_label_found = False

            for adjacent_layer_df in [prev_layer_df, next_layer_df]:
                if adjacent_layer_df is not None:
                    adj_centers = adjacent_layer_df[['Center X', 'Center Y']].values
                    if len(adj_centers) > 0:
                        distances = cdist([current_center], adj_centers)
                        min_dist_idx = np.argmin(distances)
                        min_dist = distances[0, min_dist_idx]

                        if min_dist < seuil:
                            similar_label_found = True
                            adj_label = adjacent_layer_df.iloc[min_dist_idx]['Label']
                            new_labels[labels == label] = new_labels_all_slices[z - 1][labels_all_slices[z - 1] == adj_label][0] if adjacent_layer_df is prev_layer_df else max_label + 1
                            break
            if similar_label_found:
                max_label += 1
            
            # if not similar_label_found:
            #     new_labels[labels == label] = 0
            # else:
            #     max_label += 1

        new_labels_all_slices.append(new_labels)

    return reassign_labels_sequentially(np.stack(new_labels_all_slices, axis=0))

def reassign_labels_sequentially(labels):
    unique_labels = np.unique(labels)
    unique_labels = unique_labels[unique_labels != 0]  # Exclure le label 0 s'il représente le fond
    
    # Créer un dictionnaire pour mapper les anciens labels vers les nouveaux
    label_mapping = {old_label: new_label for new_label, old_label in enumerate(unique_labels, start=1)}
    
    # Appliquer le mappage pour réassigner les labels
    new_labels = np.copy(labels)
    for old_label, new_label in label_mapping.items():
        new_labels[labels == old_label] = new_label
    
    return new_labels


"""-------------GENERAL PROCESS FUNCTIONS----------"""
def process_image(image):
    for tag_label, wavelength in tags_info[image]['tags'].items():
        process_tag(tag_label, wavelength, image)

def process_tag(tag_label, wavelength, image):
    if wavelength != "488": #488 is for the PHALLO!!
        tag_img = isolate_and_normalize_channel(wavelength, image)
        if image == "image1":
            tag_img_cropped = tag_img[up_slice_to_remove1:-down_slice_to_remove1, cropped_value:-cropped_value, cropped_value:-cropped_value]
            save_image_function(tag_img_cropped, f'{tag_label}_original.tif')
            tag_img_cropped = apply_translation_img(tag_img_cropped, image1=True)
            start_segmentation(tag_img_cropped, tag_label, 4500, 1500)
        elif image == "image2":
            tag_img_cropped = tag_img[up_slice_to_remove2:-down_slice_to_remove2, cropped_value:-cropped_value, cropped_value:-cropped_value]
            save_image_function(tag_img_cropped, f'{tag_label}_original.tif')
            tag_img_cropped = apply_translation_img(tag_img_cropped, image1=False)
            start_segmentation(tag_img_cropped, tag_label, 4500, 1500)

        save_image_function(tag_img_cropped, f'{tag_label}_cropped.tif')

    #Segmentation process
    
def process_merging():
    global circles1, circles2, scale_factor

    circles1 = find_circles('image1')
    circles2 = find_circles('image2')
    scale_factor = calcul_scale_factor()
    if scale_factor < 1:
        scale_factor = 1 / scale_factor
        apply_transfo_img(2)
    else:
        apply_transfo_img(1)

def start_segmentation(image, image_name, area_sup, area_inf):
    labels_all_slices, data = do_segmentation(image, SURFACE_THRESHOLD_SUP = area_sup, SURFACE_THRESHOLD_INF = area_inf)

    labels_all_slices = np.stack(labels_all_slices, axis=0)
    df = pd.DataFrame(data)

    save_image_function(labels_all_slices, f'label_{image_name}_non_processed.tif')
    labels = reassign_labels(labels_all_slices, df)
    save_image_function(labels, f'label_{image_name}.tif')

    tag_center[image_name] = get_normalized_center_of_labels(labels)

    return labels


"""-------------MAIN----------"""
def launch_segmentation_process():
    global json_path, project_path
    global tag_center, tags_info, model
    global count_pop, cropped_value, up_slice_to_remove1, up_slice_to_remove2, down_slice_to_remove1, down_slice_to_remove2
    global save_images

    #Global variable
    resources_path = os.environ.get('FLASK_RESOURCES_PATH', os.getcwd())
    json_path = os.path.join(resources_path, 'python', 'server', 'json') 

    #Load json data
    json_file_path = os.path.join(json_path, 'settings_data.json')
    with open(json_file_path, 'r') as f:
        tags_info = json.load(f)
    #Modify the json file (To put the process running inside the application)   
    modify_running_process("ON")
    

    project_name = tags_info["project_name"]
    project_path = os.path.join(resources_path, 'python', 'server', 'project', project_name) 
    project_path = "D:\\Users\\MFE\\Xavier\\Result_Rapport2_new"
    model = StarDist2D.from_pretrained('2D_versatile_fluo') #Load StarDist model 
    #model = StarDist2D.from_pretrained('2D_paper_dsb2018') #Load StarDist model 
    #model = StarDist3D.from_pretrained('3D_demo')
    tag_center = {}
    tag_PCA = {}
    image1 = []
    image2 = []
    count_pop = 1 
    cropped_value = tags_info["croppingValue"]
    up_slice_to_remove1 = tags_info["image1"]["upperLayersToRemove"]
    down_slice_to_remove1 = tags_info["image1"]["lowerLayersToRemove"]
    up_slice_to_remove2 = tags_info["image2"]["upperLayersToRemove"]
    down_slice_to_remove2 = tags_info["image2"]["lowerLayersToRemove"]
    save_images = tags_info["configuration"]["saveAllImages"]
    images = ["image1", "image2"]
    
    #Configuration first
    if not os.path.exists(project_path):
        os.mkdir(project_path)

    # print("HERE...")
    # struct_element = ball(5)
    # img1 = imageio.volread("C:\\Users\\MFE\\Xavier\\Outil3.0\\python\\server\\project\\ModelComp\\img_merged_pos.tif")
    # kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
    # new_img = np.zeros(img1.shape)

    # # Appliquer le closing morphologique
    # for i in range(img1.shape[0]):
    #     updt = cv2.morphologyEx(img1[i,:,:], cv2.MORPH_CLOSE, kernel)
    #     new_img[i,:,:] = cv2.morphologyEx(updt, cv2.MORPH_OPEN, kernel)
    # imageio.volwrite("C:\\Users\\MFE\\Xavier\\Outil3.0\\python\\server\\project\\ModelComp\\img_merged_pos_updt.tif",new_img )
    

    #Merge 2 images        
    print("Start merging process...")
    process_merging()
    print("Done merging")

    
    #Images Processing
    for image in images:
        process_image(image)

    #Generate csv
    print("Start generating csv...")
    generate_csv()
    print("Done generating csv")

    modify_running_process("OFF") 

#launch_segmentation_process()



