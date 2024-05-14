import json
import imageio
import czifile
import re
import cv2
import numpy as np
import pandas as pd

from xml.etree import ElementTree as ET
from csbdeep.utils import normalize
from scipy.ndimage import median_filter, center_of_mass, zoom
from skimage.transform import resize
from scipy.spatial.distance import cdist
from sklearn.decomposition import PCA
from stardist.models import StarDist2D, StarDist3D

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows


"""-------------FUNCTION----------"""
def open_image(image_path):
    with czifile.CziFile(image_path) as czi:
        image_czi = czi.asarray()
        dic_dim = dict_shape(czi)  
        axes = czi.axes  
        metadata_xml = czi.metadata()
        metadata = ET.fromstring(metadata_xml)
        #save_metadata_to_txt(metadata_xml, '/Users/xavierdekeme/Desktop/metadata.txt')
        channels_dict = channels_dict_def(metadata)

    return image_czi, dic_dim, channels_dict, axes

#return a dictionary of the czi shape dimensions:
def dict_shape(czi):
    return dict(zip([*czi.axes],czi.shape))

def channels_dict_def(metadata):
    channels = metadata.findall('.//Channel')
    channels_dict = {}
    for chan in channels:
        name = chan.attrib['Name']
        dye_element = chan.find('DyeName')
        if dye_element is not None:
            dye_numbers = re.findall(r'\d+', dye_element.text)
            dye_number = dye_numbers[-1] if dye_numbers else 'Unknown'
            channels_dict[dye_number] = name
    return channels_dict

def get_channel_index(channel_name, channel_dict):
    count = 0
    for keys, values in channel_dict.items():
        if values == channel_name:
            return count
        else:
            count += 1
  
def czi_slicer(arr,axes,indexes={"S":0, "C":0}):
    ret = arr
    axes = [*axes]
    for k,v in indexes.items():
        index = axes.index(k)
        axes.remove(k)
        ret = ret.take(axis=index, indices=v)
 
    ret_axes = ""
    for i,v in enumerate(ret.shape):
        if v > 1:
            ret_axes+=axes[i]
    return ret.squeeze(),ret_axes
 
def isolate_and_normalize_channel(image, dic_dim, channel_dict, TAG, axes, TAG_name):
    channel_name = channel_dict[TAG]
    channel_index = get_channel_index(channel_name, channel_dict)
    if channel_index < 0 or channel_index >= dic_dim['C']:
        raise ValueError("Channel index out of range.")

    image_czi_reduite, axes_restants = czi_slicer(image, axes, indexes={'C': channel_index})

    return image_czi_reduite.astype(np.uint16)

def do_segmentation(model, isolate_image, SURFACE_THRESHOLD_SUP, SURFACE_THRESHOLD_INF):
    labels_all_slices = []
    data = []

    for z in range(isolate_image.shape[0]):
        img_slice = isolate_image[z, :, :]
        img_slice = normalize(img_slice, 1, 99.8)
        filtered_img_slice = median_filter(img_slice, size=3)
        labels, details = model.predict_instances(filtered_img_slice)
        unique_labels = np.unique(labels)

        labels_to_remove = []
        for label_num in unique_labels:
            if label_num == 0: 
                continue

            instance_mask = labels == label_num
            label_surface = np.sum(instance_mask)

            if label_surface > SURFACE_THRESHOLD_SUP or label_surface < SURFACE_THRESHOLD_INF:
                labels[instance_mask] = 0
                labels_to_remove.append(label_num)

        unique_labels = np.array([label for label in unique_labels if label not in labels_to_remove])
        for label in unique_labels:
            if label == 0:  
                continue
            center = center_of_mass(labels == label)
            data.append({'Layer': z, 'Label': label, 'Center X': center[0], 'Center Y': center[1]})

        labels_all_slices.append(labels)

    return labels_all_slices, data

def reassign_labels(labels_all_slices, df, seuil=10):
    new_labels_all_slices = []
    max_label = 0  # Pour garder une trace du dernier label utilisé

    for z in range(len(labels_all_slices)):
        labels = labels_all_slices[z]
        new_labels = np.zeros_like(labels)
        layer_df = df[df['Layer'] == z]

        # Obtenir les dataframes pour les couches précédente et suivante, si elles existent
        prev_layer_df = df[df['Layer'] == z - 1] if z > 0 else None
        next_layer_df = df[df['Layer'] == z + 1] if z < len(labels_all_slices) - 1 else None

        for idx, row in layer_df.iterrows():
            label = row['Label']
            if label == 0:
                continue
            
            # Trouver les labels similaires dans les couches précédente et suivante
            current_center = np.array([row['Center X'], row['Center Y']])
            similar_label_found = False

            for adjacent_layer_df in [prev_layer_df, next_layer_df]:
                if adjacent_layer_df is not None:
                    adj_centers = adjacent_layer_df[['Center X', 'Center Y']].values
                    if len(adj_centers) > 0:
                        distances = cdist([current_center], adj_centers)
                        min_dist_idx = np.argmin(distances)
                        min_dist = distances[0, min_dist_idx]

                        if min_dist < seuil:
                            similar_label_found = True
                            adj_label = adjacent_layer_df.iloc[min_dist_idx]['Label']
                            new_labels[labels == label] = new_labels_all_slices[z - 1][labels_all_slices[z - 1] == adj_label][0] if adjacent_layer_df is prev_layer_df else max_label + 1
                            break
            if similar_label_found:
                max_label += 1
            
            # if not similar_label_found:
            #     new_labels[labels == label] = 0
            # else:
            #     max_label += 1

        new_labels_all_slices.append(new_labels)

    return reassign_labels_sequentially(np.stack(new_labels_all_slices, axis=0))

def reassign_labels_sequentially(labels):
    unique_labels = np.unique(labels)
    unique_labels = unique_labels[unique_labels != 0]  # Exclure le label 0 s'il représente le fond
    
    # Créer un dictionnaire pour mapper les anciens labels vers les nouveaux
    label_mapping = {old_label: new_label for new_label, old_label in enumerate(unique_labels, start=1)}
    
    # Appliquer le mappage pour réassigner les labels
    new_labels = np.copy(labels)
    for old_label, new_label in label_mapping.items():
        new_labels[labels == old_label] = new_label
    
    return new_labels

def project_3d_to_2d_min_layers(image_3d, min_layers=2):
    """
    Projette une image 3D en 2D en ne conservant que les pixels présents dans un minimum de couches.

    Args:
    - image_3d (numpy.ndarray): Image 3D d'entrée.
    - min_layers (int): Le nombre minimum de couches où un pixel doit être présent pour être conservé.

    Returns:
    - numpy.ndarray: Image 2D projetée.
    """
    # Initialiser une image 2D avec des zéros (background)
    image_2d = np.zeros((image_3d.shape[1], image_3d.shape[2]), dtype=image_3d.dtype)

    # Compter le nombre de fois qu'un label apparaît à chaque position (x, y)
    label_count = np.zeros_like(image_2d)

    # Itérer sur chaque couche Z de l'image 3D
    for z in range(image_3d.shape[0]):
        # Itérer sur chaque pixel (x, y) de la couche
        for y in range(image_3d.shape[1]):
            for x in range(image_3d.shape[2]):
                # Si le pixel n'est pas du fond, incrémenter le compteur
                if image_3d[z, y, x] > 0:
                    label_count[y, x] += 1

                # Si le compteur atteint le nombre minimum de couches, conserver la valeur du label
                if label_count[y, x] == min_layers:
                    image_2d[y, x] = image_3d[z, y, x]

    return image_2d
    
def find_circles_hough(image):
    image = np.uint8(image)
    circles = cv2.HoughCircles(image, cv2.HOUGH_GRADIENT, dp=2, minDist=75, param1=10, param2=90, minRadius=25, maxRadius=150)
    return circles

def calcul_scale_factor(circles1, circles2):
    radius1 = circles1[0][0][2]
    radius2 = circles2[0][0][2]
    scale_factor = radius2 / radius1 if radius1 > 0 else 1
    print(f"Scale factor: {scale_factor}")
    return scale_factor

def translate_image(image, y_offset, x_offset):
    global scale_factor

    scaled_image = zoom(image, (scale_factor, scale_factor), order=0)

    translated_image = np.zeros_like(scaled_image)
    
    for y in range(scaled_image.shape[0]):
        for x in range(scaled_image.shape[1]):
            new_y = y + y_offset
            new_x = x + x_offset
            
            # Vérifie si les nouvelles coordonnées sont dans les limites de l'image
            if 0 <= new_y < scaled_image.shape[0] and 0 <= new_x < scaled_image.shape[1]:
                translated_image[new_y, new_x] = scaled_image[y, x]
                    
    return translated_image

def merge_images_based_on_centers(circles1, circles2):
    global scale_factor

    img1 = imageio.volread(f'server/project/{project_name}/full_img1.tif')
    img2 = imageio.volread(f'server/project/{project_name}/full_img2.tif')
    center1 = circles1[0][0][:2]
    center2 = circles2[0][0][:2]

    z_dim = max(img1.shape[0], img2.shape[0])
    y_dim, x_dim = int(img1.shape[1]*scale_factor), int(img1.shape[2]*scale_factor)
    
    x_middle, y_middle = center2[0], center2[1] #utiliser le centre d'une image comme repere
    #x_middle, y_middle = x_dim //2 , y_dim//2 #utilise se le centre de l'image fictive
    x_offset_img1, x_offset_img2 = int(x_middle - center1[0] ), int(x_middle - center2[0])
    y_offset_img1, y_offset_img2 = int(y_middle - center1[1]), int(y_middle - center2[1])
    
    merged_img = np.zeros((z_dim, y_dim, x_dim), dtype=img1.dtype)
    
    mid_z_img1 = img1.shape[0] // 2
    mid_z_img2 = img2.shape[0] // 2
    mid_z_merged = z_dim // 2

    count_slice = 1
    for z in range(z_dim):
        current_slice_img1 = mid_z_img1 + z
        current_slice_img2 = mid_z_img2 + z

        if current_slice_img1 >= img1.shape[0] and current_slice_img2 >= img2.shape[0]: 
            if mid_z_img1 - count_slice >= 0:
                #slice_img1 = np.roll(np.roll(img1[mid_z_img1 - count_slice, :, :], y_offset_img1, axis=0), x_offset_img1, axis=1)
                slice_img1 = translate_image(img1[mid_z_img1 - count_slice, :, :], y_offset_img1, x_offset_img1)
            else:
                slice_img1 = np.zeros((y_dim, x_dim), dtype=img1.dtype)

            if mid_z_img2 - count_slice >= 0:
                #slice_img2 = np.roll(np.roll(img2[mid_z_img2 - count_slice, :, :], y_offset_img2, axis=0), x_offset_img2, axis=1)
                #slice_img2 = translate_image(img2[mid_z_img2 - count_slice, :, :], y_offset_img2, x_offset_img2, scale_factor)
                slice_img2 = resize(img2[mid_z_img2 - count_slice, :, :], merged_img[mid_z_merged - count_slice, :, :].shape, order=0, mode='edge', anti_aliasing=False)

            else:
                slice_img2 = np.zeros((y_dim, x_dim), dtype=img2.dtype)

            if slice_img1.shape != slice_img2.shape:
                slice_img1 = resize(slice_img1, merged_img[mid_z_merged - count_slice, :, :].shape, order=0, mode='edge', anti_aliasing=False)
                slice_img2 = resize(slice_img2, merged_img[mid_z_merged - count_slice, :, :].shape, order=0, mode='edge', anti_aliasing=False)

            merged_img[mid_z_merged - count_slice, :, :] = np.maximum.reduce([slice_img1, slice_img2])
            count_slice += 1

        else:
            if current_slice_img1 < img1.shape[0]:
                #slice_img1 = np.roll(np.roll(img1[current_slice_img1, :, :], y_offset_img1, axis=0), x_offset_img1, axis=1)
                slice_img1 = translate_image(img1[current_slice_img1, :, :], y_offset_img1, x_offset_img1)
            else:
                slice_img1 = np.zeros((y_dim, x_dim), dtype=img1.dtype)

            if current_slice_img2 < img2.shape[0]:
                #slice_img2 = np.roll(np.roll(img2[current_slice_img2, :, :], y_offset_img2, axis=0), x_offset_img2, axis=1)
                #slice_img2 = translate_image(img2[current_slice_img2, :, :], y_offset_img2, x_offset_img2, scale_factor)
                slice_img2 = resize(img2[mid_z_img2 - count_slice, :, :], merged_img[mid_z_merged + z, :, :].shape, order=0, mode='edge', anti_aliasing=False)
            else:
                slice_img2 = np.zeros((y_dim, x_dim), dtype=img2.dtype)

            if slice_img1.shape != slice_img2.shape:
                slice_img1 = resize(slice_img1, merged_img[mid_z_merged + z, :, :].shape, order=0, mode='edge', anti_aliasing=False)
                slice_img2 = resize(slice_img2, merged_img[mid_z_merged + z, :, :].shape, order=0, mode='edge', anti_aliasing=False)
                

            merged_img[mid_z_merged + z, :, :] = np.maximum.reduce([slice_img1, slice_img2])

    imageio.volwrite(f'server/project/{project_name}/img_merged_new.tif', merged_img)
    process_full_img(merged_img)
    return merged_img

def apply_translation_img(img, circles1, circles2, image1):
    global scale_factor
    center1 = circles1[0][0][:2]
    center2 = circles2[0][0][:2]

    z_dim, y_dim, x_dim = img.shape[0], int(img.shape[1]*scale_factor), int(img.shape[2]*scale_factor)
    
    x_middle, y_middle = center2[0], center2[1]
    x_offset = int(x_middle - center1[0] )
    y_offset = int(y_middle - center1[1])
    
    merged_img = np.zeros((z_dim, y_dim, x_dim), dtype=img.dtype)

    for z in range(z_dim):
        if image1:
            merged_img[z, :, :] = resize(translate_image(img[z, :, :], y_offset, x_offset), merged_img[z, :, :].shape, order=0, mode='edge', anti_aliasing=False)
        else:
            merged_img[z, :, :] = resize(img[z, :, :], merged_img[z, :, :].shape, order=0, mode='edge', anti_aliasing=False)

    return merged_img

def get_normalized_center_of_labels(label_image):
    # Trouver tous les labels uniques dans l'image (en ignorant le fond qui est généralement étiqueté par 0)
    unique_labels = np.unique(label_image)[1:]  # [1:] pour exclure 0 si le fond est inclus

    # Dimensions de l'image pour la normalisation
    dims = np.array(label_image.shape)

    # Dictionnaire pour stocker les coordonnées normalisées du centre de chaque label
    normalized_centers = {}

    # Itérer sur chaque label pour calculer et normaliser son centre de masse
    for label in unique_labels:
        # Calculer le centre de masse pour le label courant
        center = np.array(center_of_mass(label_image == label))

        # Normaliser les coordonnées du centre
        normalized_center = center / dims

        # Ajouter les coordonnées normalisées au dictionnaire
        normalized_centers[label] = normalized_center

    return normalized_centers

def get_principal_components_2D(labels_2d_image):
    """
    Applies PCA on each label in the 2D labels image to determine the principal directions along the X and Y axes.

    Args:
    - labels_2d_image (numpy.ndarray): 2D image of labels.

    Returns:
    - dict: Dictionary containing labels as keys and the principal eigenvector as values for the principal direction in 2D.
    """
    labels = np.unique(labels_2d_image)[1:] 
    principal_components = {}

    for label in labels:
        y, x = np.where(labels_2d_image == label)
        points = np.vstack((x, y)).T 

        if points.shape[0] < 2:
            continue  

        points_centered = points - np.mean(points, axis=0)

        pca = PCA(n_components=2)  
        pca.fit(points_centered)

        principal_components[label] = [pca.components_[0], pca.components_[1]]

    return principal_components

def keep_matching_regions_across_tags_PCA_dist_same_img(pop_dico, centers_TAG1, centers_TAG2, centers_TAG3, vectors_TAG1, vectors_TAG2, vectors_TAG3, distance_threshold=0.025, angle_threshold_prim=10, angle_threshold_sec=15, weights=(0, 1, 1)):
    updated_centers_TAG1, updated_centers_TAG2, updated_centers_TAG3 = {}, {}, {}
    unified_label_counter = 1

    def angle_between_vectors(v1, v2):
        dot_product = np.dot(v1, v2)
        angle = np.arccos(np.clip(dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)), -1.0, 1.0))
        return np.degrees(angle)

    def apply_weights(coordinates, weights):
        return np.multiply(coordinates, weights)

    def validate_matches(centers1, centers2, centers3, vectors1, vectors2, vectors3, distance_threshold, angle_threshold_prim, angle_threshold_sec, weights):
        nonlocal unified_label_counter
        for label1, vec1 in vectors1.items():
            match_found = False
            point1 = centers1.get(label1)
            if point1 is None: 
                continue  
            for label2, vec2 in vectors2.items():
                if match_found:
                    break
                point2 = centers2.get(label2)
                if point2 is None: 
                    continue  
                for label3, vec3 in vectors3.items():
                    point3 = centers3.get(label3)
                    if point3 is None:  # Vérifie si le point correspondant au label3 existe
                        continue  # S'il n'existe pas, passe au label3 suivant

                    angles_are_similar = True
                    # for (v1, v2) in [(vec1, vec2), (vec1, vec3), (vec2, vec3)]:
                    #     angle_prim = angle_between_vectors(v1[0], v2[0])
                    #     angle_sec = angle_between_vectors(v1[1], v2[1])
                    #     if not (angle_prim < angle_threshold_prim or abs(angle_prim - 180) < angle_threshold_prim) or not (angle_sec < angle_threshold_sec or abs(angle_sec - 180) < angle_threshold_sec):
                    #         angles_are_similar = False
                    #         break

                    if angles_are_similar:
                        point1 = centers1[label1]
                        point2 = centers2[label2]
                        point3 = centers3[label3]

                        point1_weighted = apply_weights(point1, weights)
                        point2_weighted = apply_weights(point2, weights)
                        point3_weighted = apply_weights(point3, weights)

                        diff12 = np.linalg.norm(point1_weighted - point2_weighted)
                        diff13 = np.linalg.norm(point1_weighted - point3_weighted)
                        diff23 = np.linalg.norm(point2_weighted - point3_weighted)

                        if diff12 <= distance_threshold and diff13 <= distance_threshold and diff23 <= distance_threshold:
                            unified_label = f"Unified {unified_label_counter}"
                            unified_label_counter += 1
                            for label, updated_centers in zip((label1, label2, label3), (updated_centers_TAG1, updated_centers_TAG2, updated_centers_TAG3)):
                                updated_centers[label] = {'unified_label': unified_label}

                            centers1.pop(label1)
                            centers2.pop(label2)
                            centers3.pop(label3)

                            match_found = True
                            break
            if match_found:
                continue

    validate_matches(centers_TAG1, centers_TAG2, centers_TAG3, vectors_TAG1, vectors_TAG2, vectors_TAG3, distance_threshold, angle_threshold_prim, angle_threshold_sec, weights)

    for center_dico in [updated_centers_TAG1, updated_centers_TAG2, updated_centers_TAG3]:
        pop_dico.append(center_dico)

def process_tag(image_czi, dic_dim, channels_dict, axes, tag_label, wavelength, image_list, image2 = False):
    tag_img = isolate_and_normalize_channel(image_czi, dic_dim, channels_dict, wavelength, axes, tag_label)
    if image2:
        tag_img_cropped = tag_img[6:-6, int(1024/2):, int(1024/2):]
    else:
        tag_img_cropped = tag_img[:, int(1024/2):, int(1024/2):]
    imageio.volwrite(f'server/project/{project_name}/{tag_label}_cropped.tif', tag_img_cropped)
    image_list.append(tag_img_cropped)

    #Segmentation process
    labels_all_slices, data = do_segmentation(model, tag_img_cropped, SURFACE_THRESHOLD_SUP = 700, SURFACE_THRESHOLD_INF = 75)
    labels_all_slices = np.stack(labels_all_slices, axis=0)
    df = pd.DataFrame(data)

    labels = reassign_labels(labels_all_slices, df)
    imageio.volwrite(f'server/project/{project_name}/label_{tag_label}.tif', labels)

def process_full_img(img):
    #Segmentation process
    labels_all_slices, data = do_segmentation(model, img, SURFACE_THRESHOLD_SUP = 700, SURFACE_THRESHOLD_INF = 75)
    labels_all_slices = np.stack(labels_all_slices, axis=0)
    df = pd.DataFrame(data)

    labels = reassign_labels(labels_all_slices, df)
    imageio.volwrite(f'server/project/{project_name}/label_full_img.tif', labels)

    tag_center["full_img"] = get_normalized_center_of_labels(labels)
    tag_img_2D = project_3d_to_2d_min_layers(labels)
    tag_PCA["full_img"] = get_principal_components_2D(tag_img_2D)

def get_tag_info(tag_label, image1):
    image_label = imageio.volread(f'server/project/{project_name}/label_{tag_label}.tif')
    image_label = apply_translation_img(image_label, circles1, circles2, image1) 
    imageio.volwrite(f'server/project/{project_name}/label_{tag_label}.tif', image_label)

    tag_center[tag_label] = get_normalized_center_of_labels(image_label)
    tag_img_2D = project_3d_to_2d_min_layers(image_label)
    tag_PCA[tag_label] = get_principal_components_2D(tag_img_2D)

def update_labels_image(labels_image, labels_to_keep):
    updated_labels_image = np.zeros_like(labels_image)

    for old_label, info in labels_to_keep.items():
        unified_label = int(info['unified_label'].split(" ")[-1])  # Extrait le numéro du label unifié
        #print(old_label)
        #old_label = int(old_label.split(" ")[-1])
        updated_labels_image[labels_image == old_label] = unified_label

    return updated_labels_image

def transform_label_image(label_image):
    global count_pop
    """
    Transforme les valeurs de tous les labels dans une image de label en une nouvelle valeur unique.

    Parameters:
    label_image (ndarray): Image de label d'origine.
    new_label_value (int): Nouvelle valeur à assigner à tous les labels non-zéro.

    Returns:
    ndarray: Image de label transformée.
    """
    # Créer une copie de l'image pour éviter de modifier l'originale
    transformed_image = np.copy(label_image)
    # Remplacer toutes les valeurs non-zéro par new_label_value
    transformed_image[transformed_image > 0] = count_pop
    count_pop += 1
    return transformed_image    

def process_image_list(image_list, image_number):
    composite_image = np.maximum.reduce(image_list)
    imageio.volwrite(f'server/project/{project_name}/full_img{image_number}.tif', composite_image)

    #Segmentation process
    labels_all_slices, data = do_segmentation(model, composite_image, SURFACE_THRESHOLD_SUP = 10000, SURFACE_THRESHOLD_INF = 500)
    labels_all_slices = np.stack(labels_all_slices, axis=0)
    df = pd.DataFrame(data)

    labels = reassign_labels(labels_all_slices, df)
    imageio.volwrite(f'server/project/{project_name}/label_full_img{image_number}.tif', labels)

    img_2D_full = project_3d_to_2d_min_layers(labels)
    img_circles = find_circles_hough(img_2D_full)
    return img_circles

def find_population_cell(tag_list):
    population_tag = []
    pop_list = []
    population_label = []
    
    for TAG in tag_list:
        population_tag.append(TAG[-1])
        population_label.append(imageio.volread(f'server/project/{project_name}/label_{TAG}.tif'))

    keep_matching_regions_across_tags_PCA_dist_same_img(pop_list, tag_center[tag_list[0]], tag_center[tag_list[1]], tag_center[tag_list[2]], tag_PCA[tag_list[0]], tag_PCA[tag_list[1]], tag_PCA[tag_list[2]])

    population_number = population_tag[0] + population_tag[1] + population_tag[2]
    for i, update_label in enumerate(pop_list):
        updated_labels_images = update_labels_image(population_label[i], update_label)
        #imageio.volwrite(f'/Users/xavierdekeme/Desktop/ULB-MA2/Memoire/Outil2.0/server/project/pop{population_number}/label_TAG{population_tag[i]}_pop{population_number}.tif', updated_labels_images)
        if i == 0:
            new_label_image = transform_label_image(updated_labels_images)
            imageio.volwrite(f'server/project/{project_name}/pop{population_number}.tif', new_label_image)

def is_close(center1, center2, threshold=0.0075, z_weight = 0.01):
    distance = np.sqrt(((center1[0]*z_weight - center2[0]*z_weight)**2) + (center1[1] - center2[1])**2 + (center1[2] - center2[2])**2)
    return distance <= threshold

def generate_csv(tag_center):
    tags = [tag for tag in tag_center.keys() if tag != "full_img"]
    columns = ["CellID"] + tags + ["Population"]
    data = []
    false_positive_cells = []

    for cell_id, center in enumerate(tag_center["full_img"].values(), start=1):
        row = [cell_id]
        presence_counter = 0

        for tag in tags:
            if tag in tag_center:
                match = any(is_close(center, tag_center[tag][cell]) for cell in tag_center[tag])
                presence = 1 if match else 0
                row.append(presence)
                presence_counter += presence
            else:
                row.append(0)

        # Déterminer la population en fonction de la somme des présences
        if presence_counter <= 1:
            population = "False positive"
            false_positive_cells.append(cell_id)
        elif presence_counter == 3:
            population = ''.join([tag[-1] for tag in tags if row[columns.index(tag)] == 1])
        else:
            population = "Manual Check"
        
        row.append(population)
        data.append(row)

    # Créer un DataFrame et le sauvegarder en CSV
    df = pd.DataFrame(data, columns=columns)
    df.to_csv("/Users/xavierdekeme/Desktop/Data/New/cell_presence.csv", index=False)

    # Retirer les "False positive" de l'image de label complète
    full_img_labels = imageio.volread(f'server/project/{project_name}/label_full_img.tif')
    for cell_id in false_positive_cells:
        full_img_labels[full_img_labels == cell_id] = 0

    # Sauvegarder l'image de label mise à jour
    imageio.volwrite(f'server/project/{project_name}/label_full_img_updt.tif', full_img_labels)

    # Génération du fichier Excel
    wb = Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    font = Font(bold=True)
    for col, value in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.fill = header_fill
        cell.font = font


    # Couleurs pour les différentes populations
    color_false_positive = "FFC7CE"
    color_manual_check = "FFEB9C" 
    color_population = "C6E0B4"  

    # Configuration des couleurs de fond pour les cellules de la colonne "Population"
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Appliquer la couleur de fond selon la valeur de la colonne "Population"
            if c_idx == len(columns):  # Assurez-vous que ceci correspond à l'index de la colonne "Population"
                if value == "False positive":
                    cell.fill = PatternFill(start_color=color_false_positive, end_color=color_false_positive, fill_type="solid")
                elif value == "Manual Check":
                    cell.fill = PatternFill(start_color=color_manual_check, end_color=color_manual_check, fill_type="solid")
                elif value != "":  # Toute autre valeur indiquant une population
                    cell.fill = PatternFill(start_color=color_population, end_color=color_population, fill_type="solid")

    wb.save("/Users/xavierdekeme/Desktop/Data/New/cell_presence.xlsx")



"""-------------MAIN----------"""
#Global variable
project_name = "General"
json_file_path = 'server/json/settings_data.json'
model = StarDist2D.from_pretrained('2D_versatile_fluo') #Load StarDist model 
tag_center = {}
tag_PCA = {}
image1 = []
image2 = []
count_pop = 1 

#Load json data
with open(json_file_path, 'r') as f:
    tags_info = json.load(f)

#Image1 Processing
image_czi_1, dic_dim_1, channels_dict_1, axes_1 = open_image(tags_info['image1']['path'])
for tag_label, wavelength in tags_info['image1']['tags'].items():
    process_tag(image_czi_1, dic_dim_1, channels_dict_1, axes_1, tag_label, wavelength, image1)
    print(f"Done process {tag_label}")
circles1 = process_image_list(image1, "1")

#Image2 Processing
image_czi_2, dic_dim_2, channels_dict_2, axes_2 = open_image(tags_info['image2']['path'])
for tag_label, wavelength in tags_info['image2']['tags'].items():
    process_tag(image_czi_2, dic_dim_2, channels_dict_2, axes_2, tag_label, wavelength, image2, image2=True)
    print(f"Done process {tag_label}")
circles2 = process_image_list(image2, "2")

#Merge 2 images
scale_factor = calcul_scale_factor(circles1, circles2)
merged_image = merge_images_based_on_centers(circles1, circles2)
print("Done merging")

#Get Tag info (centers, PCA, ...)
for tag_label, wavelength in tags_info['image1']['tags'].items():
    get_tag_info(tag_label, image1=True)
    print(f"Done find centers {tag_label}")

for tag_label, wavelength in tags_info['image2']['tags'].items():
    get_tag_info(tag_label, image1=False)
    print(f"Done find centers {tag_label}")

#Generate csv
generate_csv(tag_center)
print("Done generating csv")

#Find population - TO UNCOMMENT!!!!!!
# for pop, tag_list in tags_info['populations'].items():
#     find_population_cell(tag_list)